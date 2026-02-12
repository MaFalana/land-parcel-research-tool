#!/usr/bin/env python3
"""
WTHGIS Parcel Enricher - General Purpose

- Reads parcel IDs from TXT, CSV, or XLSX files
- Looks up parcels on any WTHGIS portal (configurable base URL)
- Extracts property data and downloads Property Record Card PDFs
- Outputs enriched data to a timestamped Excel file

Install:
  pip install requests beautifulsoup4 openpyxl playwright pandas
  python -m playwright install chromium

Run:
  python main.py --input parcels.txt --county "Jennings" --base-url https://jenningsin.wthgis.com
  python main.py --input parcels.csv --county "Monroe"
  python main.py --input parcels.xlsx --county "Brown" --base-url https://brownin.wthgis.com
  
  python main.py --input crawfordsville_example.txt --county "Putnam" --base-url https://putnamin.wthgis.com
  python main.py --input crawfordsville_example.txt --county "Montgomery" --base-url https://montgomeryin.wthgis.com
  python main.py --input benton_example.txt --county "Benton" --base-url https://bentonin.wthgis.com
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import random
import re
import shutil
import sys
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# BASE URL will be set via command-line argument
BASE = None


# ---------- politeness / reliability ----------
DEFAULT_PAGE_DELAY_RANGE = (2.5, 6.0)  # seconds between HTML requests
DEFAULT_PDF_DELAY_RANGE = (6.0, 12.0)  # seconds between PDF requests
DEFAULT_BROWSER_TIMEOUT_MS = 35_000


def ts() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")


def human_sleep(kind: str, page_delay_range, pdf_delay_range) -> None:
    lo, hi = page_delay_range if kind == "page" else pdf_delay_range
    time.sleep(random.uniform(lo, hi))


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w\-.]+", "_", str(s).strip())
    return s[:180] if len(s) > 180 else s


@dataclass
class ParcelData:
    soid: str
    dsid: str
    feature_id: str

    owner_name: Optional[str] = None
    owner_addr_line: Optional[str] = None
    owner_city: Optional[str] = None
    owner_state: Optional[str] = None
    owner_zip: Optional[str] = None

    situs_addr_line: Optional[str] = None
    situs_city: Optional[str] = None
    situs_state: Optional[str] = None
    situs_zip: Optional[str] = None

    legal_desc: Optional[str] = None
    document_id: Optional[str] = None  # often like 2018/3706; best-effort "latest deed/instrument"
    report_card_url: Optional[str] = None
    report_card_path: Optional[str] = None


# ---------------- Excel helpers ----------------

def find_header_row(ws: Worksheet, required_headers: List[str], search_rows: int = 10) -> int:
    """Find header row by scanning first N rows for required headers (case-insensitive)."""
    req = {h.strip().lower() for h in required_headers}
    for r in range(1, min(search_rows, ws.max_row) + 1):
        present = set()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                present.add(v.strip().lower())
        if req.issubset(present):
            return r
    raise RuntimeError(f"Could not find a header row containing: {required_headers}")


def build_header_map(ws: Worksheet, header_row: int) -> Dict[str, List[int]]:
    """
    Map header name (lower) -> list of column indices (1-based), preserving duplicates.
    Example: 'address' might map to [3, 7]
    """
    hm: Dict[str, List[int]] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            key = v.strip().lower()
            hm.setdefault(key, []).append(c)
    return hm


def col_for(hm: Dict[str, List[int]], header: str, occurrence: int = 1) -> int:
    """
    Get the column for a header. occurrence=1 means first occurrence, 2 means second, etc.
    """
    key = header.strip().lower()
    cols = hm.get(key)
    if not cols or occurrence < 1 or occurrence > len(cols):
        raise KeyError(f"Header '{header}' occurrence {occurrence} not found. Found: {cols}")
    return cols[occurrence - 1]


def ensure_columns(ws: Worksheet, hm: Dict[str, List[int]], header_row: int, new_headers: List[str]) -> Dict[str, List[int]]:
    """
    Ensure these headers exist (append to the right if missing). Returns updated map.
    Only adds if header (case-insensitive) is missing entirely (any occurrence).
    """
    existing = set(hm.keys())
    next_col = ws.max_column + 1
    for h in new_headers:
        key = h.strip().lower()
        if key not in existing:
            ws.cell(header_row, next_col).value = h
            hm.setdefault(key, []).append(next_col)
            existing.add(key)
            next_col += 1
    return hm


# ---------------- WTHGIS helpers ----------------

def extract_soid_from_url(url: str) -> Optional[str]:
    m = re.search(r"[?&]soid=(\d+)", url)
    return m.group(1) if m else None


def find_and_use_search(page, parcel_id: str) -> None:
    """
    Search for a Parcel ID using the searchBox input.
    """
    import time
    
    # Wait for page to be fully loaded
    time.sleep(2)
    
    # The search box is input#searchBox
    box = page.locator('input#searchBox')
    
    if box.count() == 0:
        raise RuntimeError("Could not locate search box (input#searchBox)")
    
    # Focus clears the placeholder text
    box.click()
    time.sleep(0.5)
    
    # Fill in the parcel ID
    box.fill(str(parcel_id).strip())
    time.sleep(0.5)
    
    # Press Enter to search
    box.press("Enter")


def lookup_dsid_featureid(parcel_id: str, base_url: str, browser_timeout_ms: int, headless: bool) -> Tuple[str, str]:
    """
    Search for a parcel and extract DSID and FeatureID from the Property Card link.
    Returns (dsid, feature_id)
    """
    import time
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context()
        page = ctx.new_page()
        page.set_default_timeout(browser_timeout_ms)

        page.goto(base_url, wait_until="domcontentloaded")
        
        find_and_use_search(page, parcel_id)

        # Wait for search results to load
        try:
            page.wait_for_function(
                "() => !document.getElementById('infoWindow').innerText.includes('Searching...')",
                timeout=browser_timeout_ms
            )
        except PlaywrightTimeoutError:
            pass
        
        time.sleep(2)
        
        # Get the Property Card link
        prop_card_link = page.locator('a:has-text("Show Property Card")').first
        
        if prop_card_link.count() == 0:
            browser.close()
            raise RuntimeError(f"Could not find Property Card link for Parcel ID '{parcel_id}'.")
        
        href = prop_card_link.get_attribute('href')
        
        # Extract DSID and FeatureID
        dsid_match = re.search(r"DSID=(\d+)", href)
        feature_match = re.search(r"FeatureID=(\d+)", href)
        
        browser.close()
        
        if not dsid_match or not feature_match:
            raise RuntimeError(f"Could not extract DSID/FeatureID from link: {href}")
        
        return dsid_match.group(1), feature_match.group(1)


def polite_get(session: requests.Session, url: str, page_delay_range, pdf_delay_range, kind: str = "page") -> requests.Response:
    human_sleep(kind, page_delay_range, pdf_delay_range)
    r = session.get(url, timeout=45)
    r.raise_for_status()
    return r


def dsid_feature_from_printpreview(session: requests.Session, soid: str, base_url: str, page_delay_range, pdf_delay_range) -> Tuple[str, str, str]:
    """
    GET printpreview1.ashx?soid=... and parse DSID + FeatureID, plus best report-card URL.
    """
    url = f"{base_url}/tgis/printpreview1.ashx?soid={soid}"
    html = polite_get(session, url, page_delay_range, pdf_delay_range, kind="page").text

    # Prefer TaxHistoryData link (has DSID & FeatureID)
    m = re.search(r"custom\.aspx\?DSID=(\d+)&FeatureID=(\d+)&RequestType=TaxHistoryData", html, re.IGNORECASE)
    if not m:
        m = re.search(r"custom\.aspx\?DSID=(\d+)&FeatureID=(\d+)&RequestType=PropertyRecordCard", html, re.IGNORECASE)
    if not m:
        raise RuntimeError(f"Could not find DSID/FeatureID for soid={soid}")

    dsid, feature_id = m.group(1), m.group(2)
    # Build report-card url (PDF)
    report_url = f"{base_url}/tgis/custom.aspx?DSID={dsid}&FeatureID={feature_id}&RequestType=PropertyRecordCard"
    return dsid, feature_id, report_url


def parse_city_state_zip(s: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Parse strings like:
      "BLOOMFIELD,IN 47424-0000"
      "SPRINGVILLE, IN 47462"
    """
    if not s:
        return None, None, None
    raw = " ".join(str(s).split())
    raw = raw.replace(", ", ",")
    m = re.search(r"^(.*?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)", raw.upper())
    if not m:
        return None, None, None
    city = m.group(1).title()
    state = m.group(2).upper()
    zipc = m.group(3)
    return city, state, zipc


def parse_parcel_info_from_search(info_html: str) -> ParcelData:
    """
    Parse parcel data directly from the search results info panel HTML.
    Handles multiple ThinkGIS formats:
    - Format 1: <th class="leftheader"> with fields like "OwnerName", "LocationAddress"
    - Format 2: <td class="ftrfld"> with fields like "mvOwnerName", "mvPropStreet"
    """
    from bs4 import BeautifulSoup
    
    soup = BeautifulSoup(info_html, "html.parser")
    
    pd = ParcelData(soid="", dsid="", feature_id="")
    
    # Build a field map from all table rows
    field_map = {}
    rows = soup.find_all('tr')
    
    for row in rows:
        # Try format 1: <th class="leftheader"> + <td>
        th = row.find('th', class_='leftheader')
        if th:
            td = row.find('td')
            if td:
                label = th.get_text(strip=True).replace('\xa0', ' ')
                value = td.get_text('\n', strip=True)
                field_map[label] = value
                continue
        
        # Try format 2: <td class="ftrfld"> + <td class="ftrval">
        tds = row.find_all('td')
        if len(tds) >= 2:
            fld = tds[0]
            val = tds[1]
            if fld.get('class') and 'ftrfld' in fld.get('class'):
                label = fld.get_text(strip=True).replace('\xa0', ' ')
                value = val.get_text('\n', strip=True)
                field_map[label] = value
    
    # Now extract data using field name variations
    # Owner Name
    pd.owner_name = (
        field_map.get('OwnerName') or 
        field_map.get('mvOwnerName') or 
        None
    )
    
    # Legal Description
    pd.legal_desc = (
        field_map.get('LegalDescription') or 
        field_map.get('mvLegalDescription') or 
        None
    )
    
    # Document/Instrument
    pd.document_id = (
        field_map.get('Document') or 
        field_map.get('mvTransferDate') or 
        None
    )
    
    # Parcel Number
    pd.parcel_number = (
        field_map.get('ParcelNumber') or 
        field_map.get('MVPParcelNumber') or 
        field_map.get('mvMVPTAXNUMBER') or 
        None
    )
    
    # Tax Account
    pd.tax_acct = (
        field_map.get('TaxAcct') or 
        field_map.get('TaxID') or 
        None
    )
    
    # Owner Address - Format 1 (combined)
    if 'OwnerAddress' in field_map:
        lines = [ln.strip() for ln in field_map['OwnerAddress'].split('\n') if ln.strip()]
        if len(lines) >= 1:
            pd.owner_addr_line = lines[0]
        if len(lines) >= 2:
            c, s, z = parse_city_state_zip(lines[1])
            pd.owner_city, pd.owner_state, pd.owner_zip = c, s, z
    
    # Owner Address - Format 2 (separate fields)
    if 'mvOwnerStreet' in field_map:
        pd.owner_addr_line = field_map['mvOwnerStreet']
    if 'mvOwnerCity' in field_map:
        pd.owner_city = field_map['mvOwnerCity']
    if 'mvOwnerState' in field_map:
        pd.owner_state = field_map['mvOwnerState']
    if 'mvOwnerZipCode' in field_map:
        pd.owner_zip = field_map['mvOwnerZipCode']
    
    # Property/Situs Address - Format 1 (combined)
    if 'LocationAddress' in field_map:
        lines = [ln.strip() for ln in field_map['LocationAddress'].split('\n') if ln.strip()]
        if len(lines) >= 1:
            pd.situs_addr_line = lines[0]
        if len(lines) >= 2:
            c, s, z = parse_city_state_zip(lines[1])
            pd.situs_city, pd.situs_state, pd.situs_zip = c, s, z
    
    # Property/Situs Address - Format 2 (separate fields)
    if 'mvPropStreet' in field_map:
        pd.situs_addr_line = field_map['mvPropStreet']
    if 'mvPropCity' in field_map:
        pd.situs_city = field_map['mvPropCity']
    if 'mvPropState' in field_map:
        pd.situs_state = field_map['mvPropState']
    if 'mvPropZipCode' in field_map:
        pd.situs_zip = field_map['mvPropZipCode']
    
    return pd
    """
    Pull TaxHistoryData (HTML-ish output) and best-effort extract OwnerName, OwnerAddress, LocationAddress, LegalDescription, Document.
    """
    url = f"{BASE}/tgis/custom.aspx?DSID={dsid}&FeatureID={feature_id}&RequestType=TaxHistoryData"
    html = polite_get(session, url, page_delay_range, pdf_delay_range, kind="page").text
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n", strip=True)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def grab_line(prefix: str) -> Optional[str]:
        p = prefix.strip()
        for ln in lines:
            if ln.startswith(p):
                rest = ln[len(p):].strip(" :|\t")
                return rest if rest else None
        return None

    def find_block(label: str) -> Tuple[Optional[str], Optional[str]]:
        """
        WTHGIS often outputs:
          OwnerAddress
          123 MAIN ST
          BLOOMFIELD,IN 47424-0000
        Or sometimes:
          OwnerAddress  123 MAIN ST
          BLOOMFIELD,IN 47424-0000
        """
        lab = label.strip()
        for i, ln in enumerate(lines):
            if ln.startswith(lab):
                remainder = ln[len(lab):].strip(" :|\t")
                if remainder:
                    street = remainder
                    csz = lines[i+1] if i+1 < len(lines) else None
                    return street, csz
                street = lines[i+1] if i+1 < len(lines) else None
                csz = lines[i+2] if i+2 < len(lines) else None
                return street, csz
        return None, None

    pd = ParcelData(soid="?", dsid=dsid, feature_id=feature_id)
    pd.owner_name = grab_line("OwnerName")
    pd.legal_desc = grab_line("LegalDescription")
    pd.document_id = grab_line("Document")

    situs_street, situs_csz = find_block("LocationAddress")
    owner_street, owner_csz = find_block("OwnerAddress")

    pd.situs_addr_line = situs_street
    if situs_csz:
        c, s, z = parse_city_state_zip(situs_csz)
        pd.situs_city, pd.situs_state, pd.situs_zip = c, s, z

    pd.owner_addr_line = owner_street
    if owner_csz:
        c, s, z = parse_city_state_zip(owner_csz)
        pd.owner_city, pd.owner_state, pd.owner_zip = c, s, z

    return pd


def download_report_card(session: requests.Session, url: str, out_path: str, page_delay_range, pdf_delay_range) -> str:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
        return out_path  # already downloaded

    r = polite_get(session, url, page_delay_range, pdf_delay_range, kind="pdf")
    # Many WTHGIS portals serve PDF with Content-Type application/pdf; but don't rely on it.
    with open(out_path, "wb") as f:
        f.write(r.content)
    return out_path


# ---------------- main pipeline ----------------


def owner_filename_stub(owner_name: str) -> str:
    """
    Returns LAST NAME or COMPANY NAME suitable for filenames.
    Strips common business suffixes to keep names shorter.
    """
    if not owner_name:
        return "UNKNOWN"

    name = owner_name.strip().upper()

    # Company / entity heuristics
    entity_keywords = [
        " LLC", " INC", " CORP", " CO", " COMPANY", " TRUST",
        " BANK", " CITY", " TOWN", " COUNTY", " SCHOOL",
        " CHURCH", " ASSOCIATION", " AUTHORITY"
    ]
    
    is_entity = any(k in name for k in entity_keywords)
    
    if is_entity:
        # Strip out common suffixes for cleaner filenames
        clean_name = name
        for suffix in [" LLC", " L.L.C.", " INC", " INC.", " CORP", " CORP.", 
                       " CO.", " CO ", " COMPANY", " TRUST", " LTD", " LTD."]:
            clean_name = clean_name.replace(suffix, "")
        
        # Also remove trailing punctuation and extra spaces
        clean_name = clean_name.strip(" .,&")
        clean_name = " ".join(clean_name.split())  # normalize spaces
        
        return safe_filename(clean_name)

    # Comma format: LAST, FIRST MI
    if "," in name:
        last = name.split(",", 1)[0].strip()
        return safe_filename(last)

    # Space-delimited personal name: take last token
    parts = name.split()
    if len(parts) > 1:
        return safe_filename(parts[-1])

    return safe_filename(name)


def read_parcel_ids_from_file(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
    """
    Read parcel IDs from TXT, CSV, or XLSX file.
    
    For TXT: reads one parcel ID per line
    For CSV: reads from first column or 'Parcel ID' column if it exists
    For XLSX: reads from first column or 'Parcel ID' column if it exists
    
    Returns list of parcel IDs (as strings, stripped of whitespace)
    """
    ext = os.path.splitext(file_path)[1].lower()
    parcel_ids = []
    
    if ext == '.txt':
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):  # Skip empty lines and comments
                    parcel_ids.append(line)
    
    elif ext == '.csv':
        df = pd.read_csv(file_path)
        # Try to find a column with 'parcel' and 'id' in the name (case-insensitive)
        parcel_col = None
        for col in df.columns:
            if 'parcel' in col.lower() and 'id' in col.lower():
                parcel_col = col
                break
        
        if parcel_col:
            parcel_ids = df[parcel_col].dropna().astype(str).str.strip().tolist()
        else:
            # Use first column
            parcel_ids = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
    
    elif ext in ['.xlsx', '.xls']:
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # Try to find a column with 'parcel' and 'id' in the name (case-insensitive)
        parcel_col = None
        for col in df.columns:
            col_str = str(col).lower()
            if 'parcel' in col_str and 'id' in col_str:
                parcel_col = col
                break
        
        if parcel_col:
            parcel_ids = df[parcel_col].dropna().astype(str).str.strip().tolist()
        else:
            # Use first column
            parcel_ids = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
    
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .txt, .csv, .xlsx, or .xls")
    
    # Remove any empty strings
    parcel_ids = [pid for pid in parcel_ids if pid]
    
    return parcel_ids


def batch_lookup_parcels(parcel_ids: List[str], base_url: str, browser_timeout_ms: int, headless: bool) -> Dict[str, Tuple[str, str, str]]:
    """
    Open browser ONCE and look up all parcels, returning a map of parcel_id -> (dsid, feature_id, info_html).
    This is much more efficient and polite than opening a browser for each parcel.
    """
    import time
    
    results = {}
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context()
        page = ctx.new_page()
        page.set_default_timeout(browser_timeout_ms)

        page.goto(base_url, wait_until="domcontentloaded")
        time.sleep(2)
        
        for idx, parcel_id in enumerate(parcel_ids, 1):
            print(f"[{idx}/{len(parcel_ids)}] Looking up {parcel_id}...")
            
            try:
                # Search
                box = page.locator('input#searchBox')
                box.click()
                time.sleep(0.3)
                box.fill("")  # Clear first
                box.fill(str(parcel_id).strip())
                time.sleep(0.3)
                box.press("Enter")
                
                # Wait for results
                try:
                    page.wait_for_function(
                        "() => !document.getElementById('infoWindow').innerText.includes('Searching...')",
                        timeout=browser_timeout_ms
                    )
                except PlaywrightTimeoutError:
                    pass
                
                time.sleep(1.5)
                
                # Get the Property Card link
                prop_card_link = page.locator('a:has-text("Show Property Card")').first
                
                if prop_card_link.count() == 0:
                    print(f"  ⚠ No Property Card link found for {parcel_id}")
                    continue
                
                href = prop_card_link.get_attribute('href')
                
                # Extract DSID and FeatureID
                dsid_match = re.search(r"DSID=(\d+)", href)
                feature_match = re.search(r"FeatureID=(\d+)", href)
                
                if dsid_match and feature_match:
                    dsid = dsid_match.group(1)
                    feature_id = feature_match.group(1)
                    
                    # Capture the info panel HTML (has all the parcel data!)
                    info_html = page.locator('#infoWindow').inner_html()
                    
                    results[parcel_id] = (dsid, feature_id, info_html)
                    print(f"  ✓ DSID={dsid}, FeatureID={feature_id}")
                else:
                    print(f"  ⚠ Could not extract DSID/FeatureID from: {href}")
                
            except Exception as e:
                print(f"  ✗ Error: {e}")
                continue
        
        browser.close()
    
    return results


def enrich_one_with_ids(parcel_id: str, dsid: str, feature_id: str, info_html: str, session: requests.Session,
                        base_url: str, page_delay_range, pdf_delay_range, downloads_dir: str) -> ParcelData:
    """
    Enrich a single parcel given its DSID, FeatureID, and info HTML (already looked up via browser).
    Uses polite delays between requests.
    """
    # Parse parcel data from the info HTML we already captured
    pd = parse_parcel_info_from_search(info_html)
    pd.dsid = dsid
    pd.feature_id = feature_id
    
    # Build Property Card PDF URL
    report_url = f"{base_url}/tgis/custom.aspx?DSID={dsid}&FeatureID={feature_id}&RequestType=PropertyRecordCard"
    pd.report_card_url = report_url

    stub = owner_filename_stub(pd.owner_name)
    fname = safe_filename(f"{stub}_{parcel_id}.pdf")
    out_path = os.path.join(downloads_dir, fname)
    pd.report_card_path = download_report_card(session, report_url, out_path, page_delay_range, pdf_delay_range)

    return pd


def main():
    ap = argparse.ArgumentParser(description="Enrich parcels from any WTHGIS portal.")
    ap.add_argument("--input", required=True, help="Input file path (TXT, CSV, or XLSX with parcel IDs).")
    ap.add_argument("--county", required=True, help="County name (e.g., 'Jennings', 'Monroe', 'Brown').")
    ap.add_argument("--base-url", help="Base URL for WTHGIS portal (e.g., https://jenningsin.wthgis.com). If not provided, will try to construct from county name.")
    ap.add_argument("--sheet", help="Worksheet name (for XLSX files only).")
    ap.add_argument("--downloads", help="Folder to store downloaded PDFs (default: PRC_<county>).")
    ap.add_argument("--output", help="Output XLSX file path (default: auto-generated with timestamp).")
    ap.add_argument("--headless", action="store_true", help="Run browser headless (default).")
    ap.add_argument("--headed", action="store_true", help="Run browser with a visible window.")
    ap.add_argument("--browser-timeout-ms", type=int, default=DEFAULT_BROWSER_TIMEOUT_MS)

    ap.add_argument("--page-delay-min", type=float, default=DEFAULT_PAGE_DELAY_RANGE[0])
    ap.add_argument("--page-delay-max", type=float, default=DEFAULT_PAGE_DELAY_RANGE[1])
    ap.add_argument("--pdf-delay-min", type=float, default=DEFAULT_PDF_DELAY_RANGE[0])
    ap.add_argument("--pdf-delay-max", type=float, default=DEFAULT_PDF_DELAY_RANGE[1])

    ap.add_argument("--max", type=int, default=0, help="Max parcels to process (0 = no limit).")
    ap.add_argument("--save-every", type=int, default=10, help="Save output workbook every N processed rows.")
    args = ap.parse_args()

    # Set base URL
    if args.base_url:
        base_url = args.base_url.rstrip('/')
    else:
        # Try to construct from county name
        county_slug = args.county.lower().replace(' ', '').replace('county', '')
        base_url = f"https://{county_slug}in.wthgis.com"
        print(f"No base URL provided, using: {base_url}")
    
    global BASE
    BASE = base_url

    headless = True
    if args.headed:
        headless = False
    elif args.headless:
        headless = True

    page_delay_range = (args.page_delay_min, args.page_delay_max)
    pdf_delay_range = (args.pdf_delay_min, args.pdf_delay_max)

    input_path = os.path.abspath(args.input)
    if not os.path.exists(input_path):
        raise FileNotFoundError(input_path)

    run_id = ts()
    county_safe = safe_filename(args.county)

    # Read parcel IDs from input file
    print(f"\n=== Reading parcel IDs from {input_path} ===")
    parcel_ids = read_parcel_ids_from_file(input_path, args.sheet)
    print(f"Found {len(parcel_ids)} parcel IDs")
    
    if not parcel_ids:
        print("No parcel IDs found in input file!")
        return 1
    
    # Apply max limit if specified
    if args.max and args.max < len(parcel_ids):
        print(f"Limiting to first {args.max} parcels")
        parcel_ids = parcel_ids[:args.max]

    # Setup output paths
    base_dir = os.path.dirname(input_path) if os.path.dirname(input_path) else '.'
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    
    if args.output:
        output_path = os.path.abspath(args.output)
    else:
        output_path = os.path.join(base_dir, f"{base_name}_{county_safe}_{run_id}_ENRICHED.xlsx")
    
    if args.downloads:
        downloads_dir = os.path.abspath(args.downloads)
    else:
        downloads_dir = os.path.join(base_dir, f"PRC_{county_safe}")
    
    os.makedirs(downloads_dir, exist_ok=True)

    # Create output workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{county_safe}County"
    
    # Create header row
    headers = [
        "Parcel ID",
        "Owner Name",
        "Owner Address",
        "Owner City",
        "Owner State",
        "Owner Zip",
        "Property Address",
        "Property City",
        "Property State",
        "Property Zip",
        "Legal Description",
        "Document/Instrument",
        "Report Card Downloaded",
        "Report Card Path",
        "Last Checked",
        "DSID",
        "FeatureID",
        "Status",
        "Notes"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx).value = header
    
    # Column indices (1-based)
    col_map = {header: idx for idx, header in enumerate(headers, 1)}

    # Session shared across requests
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; InternalParcelAudit/1.0)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive",
    })

    # STEP 1: Batch lookup all DSID/FeatureIDs using browser (ONCE)
    print(f"\n=== Looking up parcels on {base_url} ===")
    lookup_map = batch_lookup_parcels(parcel_ids, base_url, browser_timeout_ms=args.browser_timeout_ms, headless=headless)
    print(f"Successfully looked up {len(lookup_map)}/{len(parcel_ids)} parcels")
    
    # STEP 2: Politely scrape data and download PDFs
    print("\n=== Enriching parcels (polite scraping) ===")
    processed = 0
    saved = 0

    for idx, parcel_id in enumerate(parcel_ids, 1):
        row = idx + 1  # +1 because row 1 is headers
        
        # Write parcel ID
        ws.cell(row, col_map["Parcel ID"]).value = parcel_id
        
        if parcel_id not in lookup_map:
            # Mark as failed
            ws.cell(row, col_map["Last Checked"]).value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row, col_map["Status"]).value = "LOOKUP_FAILED"
            ws.cell(row, col_map["Notes"]).value = "Could not find parcel in WTHGIS search"
            continue
        
        dsid, feature_id, info_html = lookup_map[parcel_id]
        
        try:
            print(f"[{processed+1}/{len(lookup_map)}] Enriching {parcel_id}...")
            pd = enrich_one_with_ids(
                parcel_id=parcel_id,
                dsid=dsid,
                feature_id=feature_id,
                info_html=info_html,
                session=session,
                base_url=base_url,
                page_delay_range=page_delay_range,
                pdf_delay_range=pdf_delay_range,
                downloads_dir=downloads_dir,
            )

            # Write data to row
            if pd.owner_name:
                ws.cell(row, col_map["Owner Name"]).value = pd.owner_name
            if pd.owner_addr_line:
                ws.cell(row, col_map["Owner Address"]).value = pd.owner_addr_line
            if pd.owner_city:
                ws.cell(row, col_map["Owner City"]).value = pd.owner_city
            if pd.owner_state:
                ws.cell(row, col_map["Owner State"]).value = pd.owner_state
            if pd.owner_zip:
                ws.cell(row, col_map["Owner Zip"]).value = pd.owner_zip

            if pd.situs_addr_line:
                ws.cell(row, col_map["Property Address"]).value = pd.situs_addr_line
            if pd.situs_city:
                ws.cell(row, col_map["Property City"]).value = pd.situs_city
            if pd.situs_state:
                ws.cell(row, col_map["Property State"]).value = pd.situs_state
            if pd.situs_zip:
                ws.cell(row, col_map["Property Zip"]).value = pd.situs_zip

            if pd.legal_desc:
                ws.cell(row, col_map["Legal Description"]).value = pd.legal_desc
            if pd.document_id:
                ws.cell(row, col_map["Document/Instrument"]).value = pd.document_id

            ws.cell(row, col_map["Report Card Downloaded"]).value = True
            ws.cell(row, col_map["Report Card Path"]).value = pd.report_card_path
            ws.cell(row, col_map["Last Checked"]).value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row, col_map["DSID"]).value = pd.dsid
            ws.cell(row, col_map["FeatureID"]).value = pd.feature_id
            ws.cell(row, col_map["Status"]).value = "OK"
            ws.cell(row, col_map["Notes"]).value = ""

            processed += 1

        except Exception as e:
            ws.cell(row, col_map["Last Checked"]).value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row, col_map["Status"]).value = "ERROR"
            ws.cell(row, col_map["Notes"]).value = f"{type(e).__name__}: {e}"

        if processed and processed % args.save_every == 0:
            wb.save(output_path)
            saved += 1

    wb.save(output_path)

    print("\n=== DONE ===")
    print(f"County:         {args.county}")
    print(f"Base URL:       {base_url}")
    print(f"Input file:     {input_path}")
    print(f"Output saved:   {output_path}")
    print(f"PDFs saved to:  {downloads_dir}")
    print(f"Processed:      {processed}/{len(parcel_ids)} parcels")
    if saved:
        print(f"Intermediate saves: {saved}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
