import requests
import os
import shutil
import openpyxl
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import pdfplumber
import re
from datetime import datetime

BEACON_URL = 'https://beacon.schneidercorp.com/Application.aspx?AppID=327&LayerID=3469&PageTypeID=2&PageID=2293'
CURRENT_YEAR = datetime.now().year

# Create output directory if it doesn't exist
dir = 'Plainfield'  # Output Directory
if not os.path.exists(dir): # Create directory if it doesn't exist
    os.makedirs(dir)

pw = sync_playwright().start()

browser = pw.chromium.launch(headless=False)  # Set headless=True to run in the background

page = browser.new_page()
page.goto(BEACON_URL)
page.click('text=Agree') # Click agree button


def main():

    excel_file = duplicate('./Properties.xlsx') # Duplicate existing excel sheet

    dataframe = openpyxl.load_workbook(excel_file) # Define variable to load the dataframe

    dataframe1 = dataframe.active # Define variable to read sheet

    # Find column indices for data we need to update
    column_map = find_column_indices(dataframe1)
    
    # A row is an adjourner record, which starts from row 3, it spans column A to column R. From row 3 onward, parcel ID is in column 0 (A), owner name is in column 1 (B)

    col_range = dataframe1['A:R'] # Define the column range from A to R
    
    
    for row_idx, row in enumerate(dataframe1.iter_rows(min_row=3, max_row=dataframe1.max_row), start=3):
        parcel_ID = row[0].value  # Column A (0 index)
        owner_name = row[1].value  # Column B (1 index)

        # Check if Report Card is already downloaded
        if 'downloaded' in column_map:
            downloaded_col_idx = column_map['downloaded']
            if row[downloaded_col_idx].value is True:
                print(f"Skipping {parcel_ID} - Report Card already downloaded")
                continue

        if parcel_ID and owner_name:
            owner_name_clean = ''.join(c for c in owner_name if c.isalnum() or c in (' ', '_')).rstrip() # Clean up owner name to be filesystem friendly
            filename = f'{parcel_ID}_{owner_name_clean}.pdf'
            
            prc_url, legal_desc, latest_transfer_date, document_number_or_book_page = search_beacon2(parcel_ID) # Search Beacon for the parcel ID
            if prc_url:
                filepath = download(prc_url, dir, filename) # Download the PDF file
            else:
                print(f"No PRC link found for {parcel_ID}")
                continue
            
            if filepath:
                try:
                    # Parse the downloaded PDF
                    pdf_data = parse_pdf(filepath)
                    extracted_data = {
        'legal_description': legal_desc,
        'latest_deed_date': latest_transfer_date,
        'document_number': document_number_or_book_page,
        'book_page': document_number_or_book_page,
        'code': pdf_data['code']}
    
                    print(f"Parsed data for {parcel_ID}: {pdf_data}")
                    
                    # Update Excel with extracted data
                    updates = update_excel_row(dataframe1, row_idx, extracted_data, column_map)
                    print(f"Updated Excel row {row_idx}: {', '.join(updates)}")
                    
                    # Save Excel after each row
                    dataframe.save(excel_file)
                    print(f'Saved Excel file')
                except Exception as e:
                    print(f"Error processing {parcel_ID}: {e}")
                    continue
            else:
                print(f"Failed to download PDF for {parcel_ID}")
                
        else:
            print('Missing parcel ID or owner name, skipping row.')

    browser.close() # Close the browser
    print(f"\nProcessing complete! Updated Excel saved as: {excel_file}")

    
def duplicate(source):

    # Name of source file + time stamp
    output = f"{source[:-4]}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    #output = "WellsCountyIN_Owners.xlsx"

    try:
        shutil.copy(source, output)
        print(f"'{source}' successfully duplicated to '{output}'")
        return output
    except FileNotFoundError:
        print(f"Error: Source file '{source}' not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def read_excel(file):
    # Load workbook and active sheet
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # --- Handle headers (row 1: merged) ---
    headers = []
    for i, cell in enumerate(ws[1]):
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            if i == 0:
                # Special case for A1 being blank
                headers.append("Header1")
            else:
                # Fill with last non-None header (merged continuation)
                headers.append(headers[-1])

    print("Headers:", headers)

    # --- Handle sub-headers (row 2: not merged) ---
    sub_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[2]]
    print("Sub-headers:", sub_headers)

    # --- Combine headers + sub-headers for a multi-index ---
    combined_headers = [
        f"{h}_{s}" if s else h
        for h, s in zip(headers, sub_headers)
    ]
    print("Combined headers:", combined_headers)

    # --- Now read data from row 3 onward ---
    data = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        rec = dict(zip(combined_headers, row))
        data.append(rec)

    print("First few rows:", data[:3])

def search_beacon(parcel_ID):
    try:
        page.fill('input#ctlBodyPane_ctl03_ctl01_txtParcelID', parcel_ID) # Fill in the search form with the parcel ID
        
        page.press('input#ctlBodyPane_ctl03_ctl01_txtParcelID', 'Enter') # Press enter to submit the form
        
        legal_desc = page.inner_text('span#ctlBodyPane_ctl02_ctl01_lblLegalDescription') # Get Legal Description from Property Details page
    
        
        # Get the latest transfer info from the Transfers table (first row) (date, document number or book/page)

        # if exists

        date_selector = 'table#ctlBodyPane_ctl09_ctl01_gvwTransferHistory tbody tr:nth-child(1) th'
        doc_selector ='table#ctlBodyPane_ctl09_ctl01_gvwTransferHistory tbody tr:nth-child(1) td:nth-child(3)'

        if page.query_selector(date_selector) is not None:
        
            latest_transfer_date = page.inner_text('table#ctlBodyPane_ctl09_ctl01_gvwTransferHistory tbody tr:nth-child(1) th')# Get latest transfer date if exists
        else:
            latest_transfer_date = None

        if page.query_selector(doc_selector) is not None:
            document_number_or_book_page = page.inner_text('table#ctlBodyPane_ctl09_ctl01_gvwTransferHistory tbody tr:nth-child(1) td:nth-child(3)') # Get document number or book/page if exists
        else:
            document_number_or_book_page = None
        
        prc_link = page.get_attribute('a#ctlBodyPane_ctl17_ctl01_prtrFiles_ctl00_prtrFiles_Inner_ctl00_hlkName', 'href') # get latest Property Record Card link

        print(f'Property Record Card link: {prc_link}')
        print(f'Legal Description: {legal_desc}')
        print(f'Latest Transfer Date: {latest_transfer_date}')
        print(f'Document Number or Book/Page: {document_number_or_book_page}')
        page.goto(BEACON_URL)

        return prc_link, legal_desc, latest_transfer_date, document_number_or_book_page
    except Exception as e:
        print(f"Error searching Beacon for {parcel_ID}: {e}")
        page.goto(BEACON_URL)  # Try to go back to main page
        return None

def search_beacon2(parcel_ID): # Hendricks County version
    try:
        #ctlBodyPane_ctl02_ctl01_txtParcelID
        page.fill('input#ctlBodyPane_ctl02_ctl01_txtParcelID', parcel_ID) # Fill in the search form with the parcel ID
        
        page.press('input#ctlBodyPane_ctl02_ctl01_txtParcelID', 'Enter') # Press enter to submit the form
        
        legal_desc = page.inner_text('span#ctlBodyPane_ctl03_ctl01_lblLegalDescription') # Get Legal Description from Property Details page
    
        
        # Get the latest transfer info from the Transfers table (first row) (date, document number or book/page)

        # if exists

        date_selector = 'table#ctlBodyPane_ctl11_ctl01_gvwTransferHistory tbody tr:nth-child(1) th'
        doc_selector ='table#ctlBodyPane_ctl11_ctl01_gvwTransferHistory tbody tr:nth-child(1) td:nth-child(3)'

        if page.query_selector(date_selector) is not None:
        
            latest_transfer_date = page.inner_text('table#ctlBodyPane_ctl11_ctl01_gvwTransferHistory tbody tr:nth-child(1) th')# Get latest transfer date if exists
        else:
            latest_transfer_date = None

        if page.query_selector(doc_selector) is not None:
            document_number_or_book_page = page.inner_text('table#ctlBodyPane_ctl11_ctl01_gvwTransferHistory tbody tr:nth-child(1) td:nth-child(3)') # Get document number or book/page if exists
        else:
            document_number_or_book_page = None

        # get latest Property Record Card link should match the current year

        
        prc_link = page.get_attribute('a#ctlBodyPane_ctl01_ctl01_prtrFiles_ctl01_prtrFiles_Inner_ctl00_hlkName', 'href') # get latest Property Record Card link

        print(f'Property Record Card link: {prc_link}')
        print(f'Legal Description: {legal_desc}')
        print(f'Latest Transfer Date: {latest_transfer_date}')
        print(f'Document Number or Book/Page: {document_number_or_book_page}')
        page.goto(BEACON_URL)

        return prc_link, legal_desc, latest_transfer_date, document_number_or_book_page
    except Exception as e:
        print(f"Error searching Beacon for {parcel_ID}: {e}")
        page.goto(BEACON_URL)  # Try to go back to main page
        return None
               
def download(url, output, filename):
    response = requests.get(url)

    filepath = os.path.join(output, filename)

    if response.status_code == 200:
        with open(filepath, 'wb') as file:
            file.write(response.content)
        print('Successfully downloaded file: ', filename)
        return filepath
    else:
        print('Failed to download file: ', filename)
        return None

def parse_pdf(file):
    """Parse PDF to extract Legal Description and Transfer of Ownership information.
    
    Returns:
        dict: Dictionary containing:
            - legal_description: The legal description of the property
            - latest_deed_date: The most recent deed date (MM/DD/YYYY format)
            - document_number: Document number or book/page if no doc number
            - book_page: Book and page reference
            - code: Transfer code (e.g., WD, LW, QC)
    """
    
    extracted_data = {
        'legal_description': '',
        'latest_deed_date': '',
        'document_number': '',
        'book_page': '',
        'code': ''
    }
    
    full_text = ""
    
    try:
        with pdfplumber.open(file) as pdf:
            # Extract text from all pages
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
        
        # Extract Legal Description - IMPROVED
        legal_section_match = re.search(
            r'Legal\s*\n((?:[^\n]+\n?)+?)(?=Property Class|Valuation|Tax ID|Year|M\d+\.\d+|$)', 
            full_text, 
            re.IGNORECASE | re.MULTILINE
        )
        
        if legal_section_match:
            legal_section = legal_section_match.group(1)
            
            # Handle Routing Number cases more carefully
            routing_match = re.search(r'Routing Number\s+(.+?)(?:\n|$)', legal_section, re.IGNORECASE)
            
            if routing_match:
                before_routing = legal_section[:routing_match.start()].strip()
                after_routing = routing_match.group(1).strip()
                
                # Check if after_routing contains legal description keywords
                legal_keywords = r'\b(LOT|BLOCK|PARCEL|TRACT|SECTION|TOWNSHIP|ADDITION|SUBDIVISION|ORIG|ORIGINAL|PT|PART)\b'
                if re.search(legal_keywords, after_routing, re.IGNORECASE):
                    legal_desc = before_routing + ' ' + after_routing
                else:
                    legal_desc = before_routing
            else:
                legal_desc = legal_section.strip()
            
            # Clean up and format
            lines = [line.strip() for line in legal_desc.split('\n') if line.strip()]
            # Remove routing number references like "(4-C)" or "M04.14 R88"
            cleaned_lines = []
            for line in lines:
                if not re.match(r'^[\(\)A-Z0-9\.\-\s]*$', line) or len(line) > 15:
                    cleaned_lines.append(line)
            
            if cleaned_lines:
                legal_desc = ' '.join(cleaned_lines)
                extracted_data['legal_description'] = legal_desc
        
        # Extract Transfer of Ownership - IMPROVED
        transfer_section_match = re.search(
            r'Transfer of Ownership.*?\n(.*?)(?:Valuation|Legal|Local|$)', 
            full_text, 
            re.DOTALL | re.IGNORECASE
        )
        
        if transfer_section_match:
            transfer_text = transfer_section_match.group(1)
            
            # Look for structured transfer entries first
            transfer_pattern = r'(\d{1,2}/\d{1,2}/\d{4})\s+([A-Z\s,&]+?)\s+(\d{5,})\s+([A-Z]{2})\s+(\d+/\d+)'
            matches = re.findall(transfer_pattern, transfer_text)
            
            if matches:
                latest_transfer = matches[0]
                date, owner, doc_id, code, book_page = latest_transfer
                
                extracted_data['latest_deed_date'] = date
                extracted_data['document_number'] = doc_id
                extracted_data['book_page'] = book_page
                extracted_data['code'] = code.upper()
            else:
                # Fallback extraction
                # Extract dates
                date_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'
                dates_found = re.findall(date_pattern, transfer_text)
                
                if dates_found:
                    valid_dates = []
                    for date_str in dates_found:
                        try:
                            date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                            if date_obj.year > 1950:
                                valid_dates.append((date_obj, date_str))
                        except:
                            pass
                    
                    if valid_dates:
                        latest_date = max(valid_dates, key=lambda x: x[0])
                        extracted_data['latest_deed_date'] = latest_date[1]
                
                # Extract deed codes - IMPROVED
                # Look for common deed codes, being very specific about context
                code_patterns = [
                    r'\b(WD|QC|CD|LD|LW|TD|SD|GD|PD)\s*/\s*[A-Z]',  # Code followed by / and letter
                    r'\b(WD|QC|CD|LD|LW|TD|SD|GD|PD)\b(?!\s*\d)',  # Code not followed by digits
                ]
                
                for pattern in code_patterns:
                    code_match = re.search(pattern, transfer_text, re.IGNORECASE)
                    if code_match:
                        extracted_data['code'] = code_match.group(1).upper()
                        break
                
                # Extract document numbers - MUCH IMPROVED zipcode filtering
                # Known Indiana zipcodes to avoid
                known_zipcodes = {'46714', '46804', '46634', '48098', '37067', '46777', '90275', '46750', '46759', '46774', '46804'}
                
                # Find all potential document numbers
                all_numbers = re.findall(r'\b(\d{4,})\b', transfer_text)
                for num in all_numbers:
                    # Skip known zipcodes
                    if num in known_zipcodes:
                        continue
                    
                    # Skip if it appears right after state abbreviation
                    zip_after_state = rf'(?:^|\s)(?:IN|[A-Z]{{2}})\s+{re.escape(num)}\b'
                    if re.search(zip_after_state, transfer_text, re.IGNORECASE):
                        continue
                    
                    # Skip years (4 digits between 1800-2030)
                    if len(num) == 4 and 1800 <= int(num) <= 2030:
                        continue
                    
                    # Skip if it appears in address context
                    address_context = rf'\b(?:ST|STREET|AVE|AVENUE|RD|ROAD|LN|LANE|DR|DRIVE|BLVD|BOULEVARD|MAIN|MARKET|CHERRY)\s+.*{re.escape(num)}'
                    if re.search(address_context, transfer_text, re.IGNORECASE):
                        continue
                    
                    # Prefer longer numbers (6+ digits are more likely to be document numbers)
                    if len(num) >= 6:
                        extracted_data['document_number'] = num
                        break
                
                # If no 6+ digit number, try 5 digit numbers with more filtering
                if not extracted_data['document_number']:
                    for num in all_numbers:
                        if len(num) == 5 and num not in known_zipcodes:
                            # Additional context checks for 5-digit numbers
                            zip_after_state = rf'(?:^|\s)(?:IN|[A-Z]{{2}})\s+{re.escape(num)}\b'
                            if not re.search(zip_after_state, transfer_text, re.IGNORECASE):
                                extracted_data['document_number'] = num
                                break
                
                # Look for book/page if no document number
                if not extracted_data['document_number']:
                    book_page_match = re.search(r'(\d{2,})[/\s]+(\d{2,})', transfer_text)
                    if book_page_match:
                        extracted_data['book_page'] = f"{book_page_match.group(1)}/{book_page_match.group(2)}"
                        extracted_data['document_number'] = extracted_data['book_page']
    
    except Exception as e:
        print(f"Error parsing PDF {file}: {e}")
    
    return extracted_data


def find_column_indices(worksheet):
    """Find the column indices for the data we need to update.
    
    Searches row 2 (sub-headers) for column names.
    """
    column_map = {}
    
    # Check row 2 for sub-headers
    for col_idx, cell in enumerate(worksheet[2]):
        if cell.value:
            header = str(cell.value).strip().lower()
            
            # Map headers to our data fields
            if 'desc' in header and 'description' not in header:
                column_map['desc'] = col_idx
            elif 'latest deed' in header:
                column_map['latest_deed'] = col_idx
            elif 'inst' in header or 'book/page' in header:
                column_map['document_number'] = col_idx
            elif 'deed type' in header:
                column_map['code'] = col_idx
            elif 'report card' in header:
                column_map['downloaded'] = col_idx
    
    return column_map


def update_excel_row(worksheet, row_number, pdf_data, column_map):
    """Update a specific row with PDF data and apply formatting."""
    
    # Get the row
    row = list(worksheet.iter_rows(min_row=row_number, max_row=row_number))[0]
    
    updates_made = []
    
    # Create alignment style (middle vertical, center horizontal)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Update desc column
    if 'desc' in column_map and pdf_data['legal_description']:
        col_idx = column_map['desc']
        cell = row[col_idx]
        cell.value = pdf_data['legal_description']
        cell.alignment = center_alignment
        updates_made.append(f"Desc: {pdf_data['legal_description'][:50]}...")
    
    # Update latest deed column
    if 'latest_deed' in column_map and pdf_data['latest_deed_date']:
        col_idx = column_map['latest_deed']
        cell = row[col_idx]
        cell.value = pdf_data['latest_deed_date']
        cell.alignment = center_alignment
        updates_made.append(f"Latest Deed: {pdf_data['latest_deed_date']}")
    
    # Update document number column
    if 'document_number' in column_map and pdf_data['document_number']:
        col_idx = column_map['document_number']
        cell = row[col_idx]
        cell.value = pdf_data['document_number']
        cell.alignment = center_alignment
        updates_made.append(f"Document #: {pdf_data['document_number']}")
    
    # Update deed type/code column
    if 'code' in column_map and pdf_data['code']:
        col_idx = column_map['code']
        cell = row[col_idx]
        cell.value = pdf_data['code']
        cell.alignment = center_alignment
        updates_made.append(f"Deed Type: {pdf_data['code']}")
    
    # Update downloaded column (Report Card checkbox)
    if 'downloaded' in column_map:
        col_idx = column_map['downloaded']
        cell = row[col_idx]
        cell.value = True  # Boolean True for checked checkbox
        cell.alignment = center_alignment
        updates_made.append("Report Card: Checked")
    
    return updates_made


if __name__ == "__main__":
    main()