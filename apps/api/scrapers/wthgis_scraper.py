"""
WTHGIS platform scraper
Based on the source code from 1. Scrape Parcels (ThinkGIS).py
"""
from scrapers.base_scraper import BaseScraper
from typing import Dict, Callable, Optional, Tuple, List
import os
import tempfile
import time
import random
import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import asyncio
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError


# Politeness delays
DEFAULT_PAGE_DELAY_RANGE = (2.5, 6.0)  # seconds between HTML requests
DEFAULT_PDF_DELAY_RANGE = (6.0, 12.0)  # seconds between PDF requests
DEFAULT_BROWSER_TIMEOUT_MS = 35_000


def human_sleep(kind: str, page_delay_range, pdf_delay_range) -> None:
    """Sleep for a random duration to be polite to servers"""
    lo, hi = page_delay_range if kind == "page" else pdf_delay_range
    time.sleep(random.uniform(lo, hi))


def safe_filename(s: str) -> str:
    """Convert string to safe filename"""
    s = re.sub(r"[^\w\-.]+", "_", str(s).strip())
    return s[:180] if len(s) > 180 else s


def parse_city_state_zip(s: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Parse strings like:
      "BLOOMFIELD,IN 47424-0000"
      "SPRINGVILLE, IN 47462"
    """
    if not s:
        return None, None, None
    raw = " ".join(str(s).split())
    raw = raw.replace(", ", ",")
    m = re.search(r"^(.*?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)", raw.upper())
    if not m:
        return None, None, None
    city = m.group(1).title()
    state = m.group(2).upper()
    zipc = m.group(3)
    return city, state, zipc


def owner_filename_stub(owner_name: str) -> str:
    """
    Returns LAST NAME or COMPANY NAME suitable for filenames.
    Strips common business suffixes to keep names shorter.
    """
    if not owner_name:
        return "UNKNOWN"

    name = owner_name.strip().upper()

    # Company / entity heuristics
    entity_keywords = [
        " LLC", " INC", " CORP", " CO", " COMPANY", " TRUST",
        " BANK", " CITY", " TOWN", " COUNTY", " SCHOOL",
        " CHURCH", " ASSOCIATION", " AUTHORITY"
    ]
    
    is_entity = any(k in name for k in entity_keywords)
    
    if is_entity:
        # Strip out common suffixes for cleaner filenames
        clean_name = name
        for suffix in [" LLC", " L.L.C.", " INC", " INC.", " CORP", " CORP.", 
                       " CO.", " CO ", " COMPANY", " TRUST", " LTD", " LTD."]:
            clean_name = clean_name.replace(suffix, "")
        
        # Also remove trailing punctuation and extra spaces
        clean_name = clean_name.strip(" .,&")
        clean_name = " ".join(clean_name.split())  # normalize spaces
        
        return safe_filename(clean_name)

    # Comma format: LAST, FIRST MI
    if "," in name:
        last = name.split(",", 1)[0].strip()
        return safe_filename(last)

    # Space-delimited personal name: take last token
    parts = name.split()
    if len(parts) > 1:
        return safe_filename(parts[-1])

    return safe_filename(name)


def parse_parcel_info_from_search(info_html: str) -> Dict:
    """
    Parse parcel data directly from the search results info panel HTML.
    Handles multiple ThinkGIS formats:
    - Format 1: <th class="leftheader"> with fields like "OwnerName", "LocationAddress"
    - Format 2: <td class="ftrfld"> with fields like "mvOwnerName", "mvPropStreet"
    """
    soup = BeautifulSoup(info_html, "html.parser")
    
    parcel_data = {}
    
    # Build a field map from all table rows
    field_map = {}
    rows = soup.find_all('tr')
    
    for row in rows:
        # Try format 1: <th class="leftheader"> + <td>
        th = row.find('th', class_='leftheader')
        if th:
            td = row.find('td')
            if td:
                label = th.get_text(strip=True).replace('\xa0', ' ')
                value = td.get_text('\n', strip=True)
                field_map[label] = value
                continue
        
        # Try format 2: <td class="ftrfld"> + <td class="ftrval">
        tds = row.find_all('td')
        if len(tds) >= 2:
            fld = tds[0]
            val = tds[1]
            if fld.get('class') and 'ftrfld' in fld.get('class'):
                label = fld.get_text(strip=True).replace('\xa0', ' ')
                value = val.get_text('\n', strip=True)
                field_map[label] = value
    
    # Extract data using field name variations
    parcel_data['owner_name'] = (
        field_map.get('OwnerName') or 
        field_map.get('mvOwnerName') or 
        None
    )
    
    parcel_data['legal_desc'] = (
        field_map.get('LegalDescription') or 
        field_map.get('mvLegalDescription') or 
        None
    )
    
    parcel_data['document_id'] = (
        field_map.get('Document') or 
        field_map.get('mvTransferDate') or 
        None
    )
    
    # Owner Address - Format 1 (combined)
    if 'OwnerAddress' in field_map:
        lines = [ln.strip() for ln in field_map['OwnerAddress'].split('\n') if ln.strip()]
        if len(lines) >= 1:
            parcel_data['owner_addr_line'] = lines[0]
        if len(lines) >= 2:
            c, s, z = parse_city_state_zip(lines[1])
            parcel_data['owner_city'] = c
            parcel_data['owner_state'] = s
            parcel_data['owner_zip'] = z
    
    # Owner Address - Format 2 (separate fields)
    if 'mvOwnerStreet' in field_map:
        parcel_data['owner_addr_line'] = field_map['mvOwnerStreet']
    if 'mvOwnerCity' in field_map:
        parcel_data['owner_city'] = field_map['mvOwnerCity']
    if 'mvOwnerState' in field_map:
        parcel_data['owner_state'] = field_map['mvOwnerState']
    if 'mvOwnerZipCode' in field_map:
        parcel_data['owner_zip'] = field_map['mvOwnerZipCode']
    
    # Property/Situs Address - Format 1 (combined)
    if 'LocationAddress' in field_map:
        lines = [ln.strip() for ln in field_map['LocationAddress'].split('\n') if ln.strip()]
        if len(lines) >= 1:
            parcel_data['situs_addr_line'] = lines[0]
        if len(lines) >= 2:
            c, s, z = parse_city_state_zip(lines[1])
            parcel_data['situs_city'] = c
            parcel_data['situs_state'] = s
            parcel_data['situs_zip'] = z
    
    # Property/Situs Address - Format 2 (separate fields)
    if 'mvPropStreet' in field_map:
        parcel_data['situs_addr_line'] = field_map['mvPropStreet']
    if 'mvPropCity' in field_map:
        parcel_data['situs_city'] = field_map['mvPropCity']
    if 'mvPropState' in field_map:
        parcel_data['situs_state'] = field_map['mvPropState']
    if 'mvPropZipCode' in field_map:
        parcel_data['situs_zip'] = field_map['mvPropZipCode']
    
    return parcel_data


async def batch_lookup_parcels_async(parcel_ids: List[str], base_url: str, browser_timeout_ms: int = DEFAULT_BROWSER_TIMEOUT_MS, headless: bool = True) -> Dict[str, Tuple[str, str, str]]:
    """
    Open browser ONCE and look up all parcels, returning a map of parcel_id -> (dsid, feature_id, info_html).
    This is much more efficient and polite than opening a browser for each parcel.
    
    Async version to work properly on Windows in background threads.
    """
    results = {}
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        ctx = await browser.new_context()
        page = await ctx.new_page()
        page.set_default_timeout(browser_timeout_ms)

        await page.goto(base_url, wait_until="domcontentloaded")
        await asyncio.sleep(2)
        
        for idx, parcel_id in enumerate(parcel_ids, 1):
            print(f"[{idx}/{len(parcel_ids)}] Looking up {parcel_id}...")
            
            try:
                # Search
                box = page.locator('input#searchBox')
                await box.click()
                await asyncio.sleep(0.3)
                await box.fill("")  # Clear first
                await box.fill(str(parcel_id).strip())
                await asyncio.sleep(0.3)
                await box.press("Enter")
                
                # Wait for results
                try:
                    await page.wait_for_function(
                        "() => !document.getElementById('infoWindow').innerText.includes('Searching...')",
                        timeout=browser_timeout_ms
                    )
                except PlaywrightTimeoutError:
                    pass
                
                await asyncio.sleep(1.5)
                
                # Get the Property Card link
                prop_card_link = page.locator('a:has-text("Show Property Card")').first
                
                if await prop_card_link.count() == 0:
                    print(f"  ⚠ No Property Card link found for {parcel_id}")
                    continue
                
                href = await prop_card_link.get_attribute('href')
                
                # Extract DSID and FeatureID
                dsid_match = re.search(r"DSID=(\d+)", href)
                feature_match = re.search(r"FeatureID=(\d+)", href)
                
                if dsid_match and feature_match:
                    dsid = dsid_match.group(1)
                    feature_id = feature_match.group(1)
                    
                    # Capture the info panel HTML (has all the parcel data!)
                    info_html = await page.locator('#infoWindow').inner_html()
                    
                    results[parcel_id] = (dsid, feature_id, info_html)
                    print(f"  ✓ DSID={dsid}, FeatureID={feature_id}")
                else:
                    print(f"  ⚠ Could not extract DSID/FeatureID from: {href}")
                
            except Exception as e:
                print(f"  ✗ Error: {e}")
                continue
        
        await browser.close()
    
    return results


def batch_lookup_parcels(parcel_ids: List[str], base_url: str, browser_timeout_ms: int = DEFAULT_BROWSER_TIMEOUT_MS, headless: bool = True) -> Dict[str, Tuple[str, str, str]]:
    """
    Synchronous wrapper for batch_lookup_parcels_async.
    
    On Windows, uses subprocess to avoid asyncio threading issues.
    On Linux, uses async approach directly.
    """
    import sys
    import subprocess
    import json
    import os
    
    # On Windows, use subprocess approach to avoid asyncio issues
    if sys.platform == 'win32':
        print("Using subprocess approach for Windows compatibility...")
        
        # Prepare input data
        input_data = {
            "parcel_ids": parcel_ids,
            "base_url": base_url,
            "browser_timeout_ms": browser_timeout_ms
        }
        
        # Get path to subprocess script
        script_path = os.path.join(os.path.dirname(__file__), "wthgis_scraper_subprocess.py")
        
        print(f"Launching subprocess: {script_path}")
        print(f"Processing {len(parcel_ids)} parcels...")
        
        # Run subprocess with longer timeout (5 minutes for 100 parcels)
        timeout_seconds = max(300, len(parcel_ids) * 3)  # 3 seconds per parcel minimum
        
        try:
            result = subprocess.run(
                [sys.executable, script_path],
                input=json.dumps(input_data),
                capture_output=True,
                text=True,
                timeout=timeout_seconds
            )
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"Subprocess timed out after {timeout_seconds} seconds")
        
        if result.returncode != 0:
            print(f"Subprocess stderr: {result.stderr}")
            raise RuntimeError(f"Subprocess failed with code {result.returncode}: {result.stderr}")
        
        print(f"Subprocess completed successfully")
        
        # Parse results
        try:
            results_dict = json.loads(result.stdout)
        except json.JSONDecodeError as e:
            print(f"Failed to parse subprocess output: {result.stdout[:500]}")
            raise RuntimeError(f"Invalid JSON from subprocess: {e}")
        
        # Convert to expected format
        results = {}
        for parcel_id, data in results_dict.items():
            results[parcel_id] = (data["dsid"], data["feature_id"], data["info_html"])
        
        print(f"Successfully looked up {len(results)}/{len(parcel_ids)} parcels")
        return results
    
    else:
        # On Linux, use async approach
        if sys.platform == 'win32':
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(
                batch_lookup_parcels_async(parcel_ids, base_url, browser_timeout_ms, headless)
            )
        finally:
            loop.close()


def download_report_card(session: requests.Session, url: str, out_path: str, page_delay_range, pdf_delay_range) -> str:
    """Download Property Record Card PDF with polite delays"""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
        return out_path  # already downloaded

    human_sleep("pdf", page_delay_range, pdf_delay_range)
    r = session.get(url, timeout=45)
    r.raise_for_status()
    
    with open(out_path, "wb") as f:
        f.write(r.content)
    return out_path


class WTHGISScraper(BaseScraper):
    """Scraper for WTHGIS (ThinkGIS) platform"""
    
    def __init__(self):
        super().__init__()
        self.platform_name = "WTHGIS"
    
    def scrape_parcels(
        self,
        parcel_file_path: str,
        base_url: str,
        county: str,
        job_id: str,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> Dict:
        """
        Scrape parcel data from WTHGIS portal
        
        Steps:
        1. Read parcel IDs from file
        2. Batch lookup all parcels using Playwright (open browser once)
        3. Extract DSID, FeatureID, and info HTML for each parcel
        4. Politely scrape detailed data and download PDFs
        5. Generate enriched Excel file
        """
        # Read parcel IDs
        parcel_ids = self.read_parcel_ids(parcel_file_path)
        total_parcels = len(parcel_ids)
        
        print(f"WTHGIS Scraper: Processing {total_parcels} parcels for {county} county")
        
        # Create output directories
        output_dir = os.path.join(tempfile.gettempdir(), "parcel_jobs", job_id, "output")
        pdfs_dir = os.path.join(output_dir, "property_cards")
        os.makedirs(pdfs_dir, exist_ok=True)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{county}_Parcels"
        
        # Create headers
        headers = [
            "Parcel ID", "Owner Name", "Owner Address", "Owner City", 
            "Owner State", "Owner Zip", "Property Address", "Property City",
            "Property State", "Property Zip", "Legal Description",
            "Document/Instrument", "Report Card Path", "Status", "Notes"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(1, col_idx).value = header
        
        # Column map
        col_map = {header: idx for idx, header in enumerate(headers, 1)}
        
        # STEP 1: Batch lookup all DSID/FeatureIDs using browser (ONCE)
        print(f"\n=== Looking up parcels on {base_url} ===")
        lookup_map = batch_lookup_parcels(parcel_ids, base_url, headless=True)
        print(f"Successfully looked up {len(lookup_map)}/{total_parcels} parcels")
        
        # STEP 2: Process each parcel and download PDFs
        print("\n=== Enriching parcels (polite scraping) ===")
        
        # Session for PDF downloads
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (compatible; InternalParcelAudit/1.0)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        })
        
        page_delay_range = DEFAULT_PAGE_DELAY_RANGE
        pdf_delay_range = DEFAULT_PDF_DELAY_RANGE
        
        processed = 0
        failed = 0
        
        for idx, parcel_id in enumerate(parcel_ids, 1):
            row = idx + 1  # +1 because row 1 is headers
            
            # Write parcel ID
            ws.cell(row, col_map["Parcel ID"]).value = parcel_id
            
            if parcel_id not in lookup_map:
                # Mark as failed
                ws.cell(row, col_map["Status"]).value = "LOOKUP_FAILED"
                ws.cell(row, col_map["Notes"]).value = "Could not find parcel in WTHGIS search"
                failed += 1
                if progress_callback:
                    progress_callback(processed, total_parcels)
                continue
            
            dsid, feature_id, info_html = lookup_map[parcel_id]
            
            try:
                print(f"[{processed+1}/{len(lookup_map)}] Enriching {parcel_id}...")
                
                # Parse parcel data from info HTML
                parcel_data = parse_parcel_info_from_search(info_html)
                
                # Write data to Excel
                ws.cell(row, col_map["Owner Name"]).value = parcel_data.get('owner_name')
                ws.cell(row, col_map["Owner Address"]).value = parcel_data.get('owner_addr_line')
                ws.cell(row, col_map["Owner City"]).value = parcel_data.get('owner_city')
                ws.cell(row, col_map["Owner State"]).value = parcel_data.get('owner_state')
                ws.cell(row, col_map["Owner Zip"]).value = parcel_data.get('owner_zip')
                
                ws.cell(row, col_map["Property Address"]).value = parcel_data.get('situs_addr_line')
                ws.cell(row, col_map["Property City"]).value = parcel_data.get('situs_city')
                ws.cell(row, col_map["Property State"]).value = parcel_data.get('situs_state')
                ws.cell(row, col_map["Property Zip"]).value = parcel_data.get('situs_zip')
                
                ws.cell(row, col_map["Legal Description"]).value = parcel_data.get('legal_desc')
                ws.cell(row, col_map["Document/Instrument"]).value = parcel_data.get('document_id')
                
                # Download Property Record Card PDF
                report_url = f"{base_url}/tgis/custom.aspx?DSID={dsid}&FeatureID={feature_id}&RequestType=PropertyRecordCard"
                
                stub = owner_filename_stub(parcel_data.get('owner_name'))
                fname = safe_filename(f"{stub}_{parcel_id}.pdf")
                pdf_path = os.path.join(pdfs_dir, fname)
                
                download_report_card(session, report_url, pdf_path, page_delay_range, pdf_delay_range)
                ws.cell(row, col_map["Report Card Path"]).value = pdf_path
                
                ws.cell(row, col_map["Status"]).value = "SUCCESS"
                processed += 1
                
            except Exception as e:
                print(f"  ✗ Error processing {parcel_id}: {e}")
                ws.cell(row, col_map["Status"]).value = "FAILED"
                ws.cell(row, col_map["Notes"]).value = str(e)
                failed += 1
            
            # Update progress
            if progress_callback:
                progress_callback(processed, total_parcels)
            
            # Save every 10 parcels
            if idx % 10 == 0:
                excel_path = os.path.join(output_dir, f"{county}_parcels_enriched.xlsx")
                wb.save(excel_path)
                print(f"  💾 Saved progress ({idx}/{total_parcels})")
        
        # Final save
        excel_path = os.path.join(output_dir, f"{county}_parcels_enriched.xlsx")
        wb.save(excel_path)
        
        print(f"\nWTHGIS Scraper: Complete!")
        print(f"  Processed: {processed}/{total_parcels}")
        print(f"  Failed: {failed}")
        print(f"  Excel: {excel_path}")
        print(f"  PDFs: {pdfs_dir}")
        
        return {
            "excel_path": excel_path,
            "pdfs_dir": pdfs_dir,
            "parcel_count": total_parcels,
            "failed_count": failed
        }
