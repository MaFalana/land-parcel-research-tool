"""
Hamilton County (ArcGIS Experience) scraper

Handles scraping from Hamilton County's ArcGIS Experience portal.
This is a JavaScript-heavy site that requires Playwright for interaction.

Workflow:
1. Navigate to ArcGIS Experience page
2. Search for parcel by Parcel ID
3. Click on search result to open popup
4. Navigate to property information page (external link)
5. Extract owner data from Ownership Information tab
6. Navigate to Property Assessment tab
7. Download most recent Property Record Card PDF
"""
from scrapers.base_scraper import BaseScraper
from typing import Dict, Callable, Optional
import os
import tempfile
import time
import random
import re
import requests
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class HamiltonScraper(BaseScraper):
    """Scraper for Hamilton County ArcGIS Experience platform"""
    
    def __init__(self):
        super().__init__()
        self.platform_name = "Hamilton County ArcGIS"
        # Delay ranges (in seconds)
        self.page_delay_range = (2, 4)  # 2-4 seconds between page requests
        self.pdf_delay_range = (1, 3)   # 1-3 seconds before PDF downloads
    
    def scrape_parcels(
        self,
        parcel_file_path: str,
        base_url: str,
        county: str,
        job_id: str,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> Dict:
        """
        Scrape parcel data from Hamilton County ArcGIS Experience portal
        
        Steps:
        1. Parse parcel IDs from input file
        2. Open browser and navigate to ArcGIS Experience
        3. For each parcel: search, extract data, download PRC
        4. Save to Excel with progress tracking
        
        Args:
            parcel_file_path: Path to file with parcel IDs
            base_url: ArcGIS Experience URL
            county: County name
            job_id: Job ID for output files
            progress_callback: Function to report progress
            
        Returns:
            Dict with excel_path and stats
        """
        # Parse parcel IDs
        parcel_ids = self.read_parcel_ids(parcel_file_path)
        total_parcels = len(parcel_ids)
        
        print(f"Hamilton Scraper: Processing {total_parcels} parcels for {county} county")
        
        # Create output directories
        output_dir = os.path.join(tempfile.gettempdir(), "parcel_jobs", job_id, "output")
        pdfs_dir = os.path.join(output_dir, "property_cards")
        os.makedirs(pdfs_dir, exist_ok=True)
        
        # Create Excel workbook
        excel_path = os.path.join(output_dir, f"{county}_hamilton_data.xlsx")
        wb = self._create_excel_template(county)
        ws = wb.active
        
        # Track progress
        processed = 0
        failed = 0
        
        # Create requests session for PDF downloads
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/pdf,*/*",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        })
        
        # Launch browser
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox'
                ]
            )
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            try:
                # We navigate directly to property pages, no need for ArcGIS Experience
                print(f"Browser ready for direct property page access")
                
                # Process each parcel
                for idx, parcel_id in enumerate(parcel_ids, start=1):
                    row_num = idx + 2  # Data starts at row 3 (1-indexed)
                    
                    try:
                        print(f"Processing {idx}/{total_parcels}: {parcel_id}")
                        
                        # Search for parcel and extract data
                        parcel_data = self._search_parcel(page, parcel_id, base_url)
                        
                        if parcel_data:
                            # Write to Excel
                            ws.cell(row_num, 1, parcel_id)
                            ws.cell(row_num, 2, parcel_data.get('state_parcel_no', ''))
                            ws.cell(row_num, 3, parcel_data.get('owner_name', ''))
                            ws.cell(row_num, 4, parcel_data.get('owner_address', ''))
                            ws.cell(row_num, 5, parcel_data.get('owner_city', ''))
                            ws.cell(row_num, 6, parcel_data.get('owner_state', ''))
                            ws.cell(row_num, 7, parcel_data.get('owner_zip', ''))
                            ws.cell(row_num, 8, parcel_data.get('parcel_address', ''))
                            ws.cell(row_num, 9, parcel_data.get('parcel_city', ''))
                            ws.cell(row_num, 10, parcel_data.get('parcel_state', ''))
                            ws.cell(row_num, 11, parcel_data.get('parcel_zip', ''))
                            ws.cell(row_num, 12, parcel_data.get('legal_description', ''))
                            ws.cell(row_num, 13, parcel_data.get('latest_deed_date', ''))
                            ws.cell(row_num, 14, parcel_data.get('document_number', ''))
                            ws.cell(row_num, 15, parcel_data.get('deed_code', ''))
                            
                            # Download PRC PDF if available
                            prc_path = ''
                            if parcel_data.get('prc_url'):
                                try:
                                    owner_stub = self._owner_filename_stub(parcel_data.get('owner_name', 'Unknown'))
                                    pdf_filename = self._safe_filename(f"{parcel_id}_{owner_stub}.pdf")
                                    prc_full_path = os.path.join(pdfs_dir, pdf_filename)
                                    
                                    self._download_prc(session, parcel_data['prc_url'], prc_full_path)
                                    prc_path = prc_full_path
                                    print(f"  ✓ Downloaded PRC: {pdf_filename}")
                                except Exception as e:
                                    print(f"  ✗ Failed to download PRC: {e}")
                                    prc_path = f"ERROR: {str(e)[:50]}"
                            
                            ws.cell(row_num, 16, prc_path)
                            ws.cell(row_num, 17, 'SUCCESS')
                            
                            processed += 1
                        else:
                            ws.cell(row_num, 1, parcel_id)
                            ws.cell(row_num, 17, 'NOT_FOUND')
                            failed += 1
                        
                        # Save progress every 10 parcels
                        if idx % 10 == 0:
                            wb.save(excel_path)
                            print(f"Progress saved: {processed} successful, {failed} failed")
                        
                        # Report progress
                        if progress_callback:
                            progress_callback(idx, total_parcels)
                        
                        # Polite delay between parcels
                        delay = random.uniform(*self.page_delay_range)
                        time.sleep(delay)
                        
                    except Exception as e:
                        print(f"Error processing {parcel_id}: {e}")
                        import traceback
                        traceback.print_exc()
                        ws.cell(row_num, 1, parcel_id)
                        ws.cell(row_num, 17, f'ERROR: {str(e)[:50]}')
                        failed += 1
                        continue
                
            finally:
                browser.close()
        
        # Final save
        wb.save(excel_path)
        
        print(f"\nHamilton Scraper: Complete!")
        print(f"  Processed: {processed}/{total_parcels}")
        print(f"  Failed: {failed}")
        print(f"  Excel: {excel_path}")
        print(f"  PDFs: {pdfs_dir}")
        
        return {
            "excel_path": excel_path,
            "pdfs_dir": pdfs_dir,
            "total": total_parcels,
            "processed": processed,
            "failed": failed
        }
    
    def _search_parcel(self, page, parcel_id: str, base_url: str) -> Optional[Dict]:
        """
        Search for a parcel and extract data
        
        Hamilton County property pages can be accessed directly via URL:
        https://secure2.hamiltoncounty.in.gov/propertyreports/reports.aspx?parcel={parcel_id_no_dashes}
        
        This bypasses the need for ArcGIS search and map interaction.
        
        Args:
            page: Playwright page object
            parcel_id: Parcel ID to search for (e.g., "03-06-06-00-00-022.000")
            base_url: Base URL of ArcGIS Experience (not used, kept for compatibility)
        
        Returns dict with parcel data or None if not found
        """
        try:
            print(f"  Accessing property page for: {parcel_id}")
            
            # Convert parcel ID to URL format (remove dashes and dots)
            # Example: "03-06-06-00-00-022.000" -> "0306060000022000"
            parcel_id_clean = parcel_id.replace('-', '').replace('.', '')
            
            # Construct direct URL to property page
            property_url = f"https://secure2.hamiltoncounty.in.gov/propertyreports/reports.aspx?parcel={parcel_id_clean}"
            
            print(f"  Navigating to: {property_url}")
            
            # Navigate directly to property page
            response = page.goto(property_url, wait_until="domcontentloaded", timeout=60000)
            
            # Check if page loaded successfully
            if response.status != 200:
                print(f"  ✗ Property page returned status {response.status}")
                return None
            
            # Wait for page to fully load - Hamilton County pages can be slow
            page.wait_for_timeout(5000)
            
            # Check if we got a valid property page (look for owner info)
            try:
                owner_elem = page.locator('span#ownerCR').first
                owner_elem.wait_for(state="attached", timeout=10000)
                print(f"  ✓ Property page loaded successfully")
            except Exception as e:
                print(f"  ✗ Property not found or page did not load correctly: {e}")
                # Save HTML for debugging
                try:
                    page_title = page.title()
                    print(f"  Page title: {page_title}")
                except:
                    pass
                return None
            
            # Extract data from property page
            data = self._extract_property_data(page, parcel_id)
            
            if not data:
                print(f"  ✗ Could not extract property data")
                return None
            
            return data
            
        except Exception as e:
            print(f"Error accessing property page for {parcel_id}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_property_data(self, page, parcel_id: str) -> Optional[Dict]:
        """
        Extract property data from Hamilton County property information page
        
        Hamilton County uses specific span IDs for data:
        - ownerCR: Owner name
        - ownerAddress1CR, ownerAddress2CR, ownerAddress3CR: Mailing address
        - Property location is in the page text
        - PRC links are in the Assessment section
        
        Returns dict with all property fields
        """
        data = {
            'state_parcel_no': parcel_id,  # Hamilton uses same format
            'owner_name': '',
            'owner_address': '',
            'owner_city': '',
            'owner_state': '',
            'owner_zip': '',
            'parcel_address': '',
            'parcel_city': '',
            'parcel_state': '',
            'parcel_zip': '',
            'legal_description': '',
            'latest_deed_date': '',
            'document_number': '',
            'deed_code': '',
            'prc_url': None
        }
        
        try:
            print(f"  Extracting property data...")
            
            # Extract Owner Name from span#ownerCR
            try:
                # Wait for the page to fully load
                page.wait_for_load_state('networkidle', timeout=10000)
                page.wait_for_timeout(2000)
                
                owner_elem = page.locator('span#ownerCR').first
                owner_elem.wait_for(state="attached", timeout=5000)
                data['owner_name'] = owner_elem.inner_text().strip()
                print(f"    Owner: {data['owner_name']}")
            except Exception as e:
                print(f"    Could not find owner name with span#ownerCR: {e}")
                # Try alternative: look in page HTML directly
                try:
                    page_html = page.content()
                    import re
                    owner_match = re.search(r'<span id="ownerCR">(.*?)</span>', page_html)
                    if owner_match:
                        data['owner_name'] = owner_match.group(1).strip()
                        print(f"    Owner (from HTML): {data['owner_name']}")
                except Exception as e2:
                    print(f"    Could not extract owner from HTML: {e2}")
            
            # Extract Owner Address from span#ownerAddress1CR, ownerAddress2CR, ownerAddress3CR
            try:
                addr_parts = []
                page_html = page.content()
                import re
                
                # Extract from HTML directly for reliability
                for addr_id in ['ownerAddress1CR', 'ownerAddress2CR', 'ownerAddress3CR']:
                    match = re.search(rf'<span id="{addr_id}">(.*?)</span>', page_html)
                    if match:
                        text = match.group(1).strip()
                        if text and text != ' ':
                            addr_parts.append(text)
                
                if addr_parts:
                    # Combine all address parts
                    full_address = ', '.join(addr_parts)
                    # Parse into components
                    parsed = self._parse_address(full_address)
                    data['owner_address'] = parsed['street']
                    data['owner_city'] = parsed['city']
                    data['owner_state'] = parsed['state']
                    data['owner_zip'] = parsed['zip']
                    print(f"    Mailing Address: {full_address}")
            except Exception as e:
                print(f"    Could not extract owner address: {e}")
            
            # Extract Property Location
            try:
                # Look for "Property Location:" in the page text
                page_text = page.content()
                import re
                prop_loc_match = re.search(r'Property Location:\s*([^<]+)', page_text)
                if prop_loc_match:
                    prop_location = prop_loc_match.group(1).strip()
                    parsed = self._parse_address(prop_location)
                    
                    # Skip street number if it's "0"
                    street = parsed['street']
                    if street.startswith('0 '):
                        street = street[2:]  # Remove "0 " prefix
                    
                    data['parcel_address'] = street
                    data['parcel_city'] = parsed['city']
                    data['parcel_state'] = parsed['state']
                    data['parcel_zip'] = parsed['zip']
                    print(f"    Property Location: {prop_location}")
            except Exception as e:
                print(f"    Could not extract property location: {e}")
            
            # Try to click Ownership tab to get more details
            try:
                ownership_tab = page.locator('a#ownershipInfo, button:has-text("Ownership")').first
                if ownership_tab.is_visible(timeout=2000):
                    ownership_tab.click()
                    page.wait_for_timeout(1000)
                    
                    # Look for legal description
                    try:
                        legal_elem = page.locator('text=/Legal Description/i').first
                        if legal_elem.is_visible(timeout=2000):
                            # Get the next element or parent that contains the actual description
                            parent = legal_elem.locator('xpath=..').first
                            legal_text = parent.inner_text()
                            # Extract just the description part
                            import re
                            legal_match = re.search(r'Legal Description[:\s]+(.*)', legal_text, re.IGNORECASE | re.DOTALL)
                            if legal_match:
                                data['legal_description'] = legal_match.group(1).strip()[:500]  # Limit length
                    except:
                        pass
                    
                    # Look for deed information in transfer table
                    try:
                        # Hamilton County has a transfer table with deed info
                        transfer_rows = page.locator('tbody#transferTable tr').all()
                        if transfer_rows:
                            # Get the first (most recent) transfer
                            first_row = transfer_rows[0]
                            row_text = first_row.inner_text()
                            
                            # Extract date (first cell)
                            cells = first_row.locator('td').all()
                            if cells:
                                date_text = cells[0].inner_text().strip()
                                # Extract just the date part (before any newline)
                                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', date_text)
                                if date_match:
                                    data['latest_deed_date'] = date_match.group(1)
                    except:
                        pass
            except:
                pass
            
            # Now look for Property Record Card in Assessment section
            print(f"  Looking for Property Record Card...")
            
            try:
                # Try to find and click Assessment tab
                assessment_selectors = [
                    'a#propertyAssessments',
                    'button:has-text("Property Assessments")',
                    'a:has-text("Property Assessments")'
                ]
                
                for selector in assessment_selectors:
                    try:
                        assessment_tab = page.locator(selector).first
                        if assessment_tab.is_visible(timeout=2000):
                            assessment_tab.click()
                            page.wait_for_timeout(2000)
                            break
                    except:
                        continue
            except:
                pass
            
            # Look for PRC links with pattern: /publicdocs/PRC/PRC{year}/{id}.pdf
            try:
                page_html = page.content()
                import re
                
                # Find all PRC links
                prc_pattern = r'href="(https://secure2\.hamiltoncounty\.in\.gov/publicdocs/PRC/PRC(\d{4})/\d+\.pdf)"'
                prc_matches = re.findall(prc_pattern, page_html)
                
                if prc_matches:
                    print(f"    Found {len(prc_matches)} Property Record Cards")
                    
                    # Select the most recent year
                    best_url = None
                    best_year = 0
                    
                    for url, year in prc_matches:
                        year_int = int(year)
                        if year_int > best_year:
                            best_year = year_int
                            best_url = url
                    
                    if best_url:
                        data['prc_url'] = best_url
                        print(f"    Selected PRC from year {best_year}")
                else:
                    print(f"    No Property Record Cards found")
                    
            except Exception as e:
                print(f"    Error finding PRC: {e}")
            
            return data
            
        except Exception as e:
            print(f"  Error extracting property data: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_popup_data(self, page, parcel_id: str) -> Optional[Dict]:
        """
        Extract property data from ArcGIS popup (if property page not available)
        
        Returns dict with available fields from popup
        """
        data = {
            'state_parcel_no': '',
            'owner_name': '',
            'owner_address': '',
            'owner_city': '',
            'owner_state': '',
            'owner_zip': '',
            'parcel_address': '',
            'parcel_city': '',
            'parcel_state': '',
            'parcel_zip': '',
            'legal_description': '',
            'latest_deed_date': '',
            'document_number': '',
            'deed_code': '',
            'prc_url': None
        }
        
        try:
            # Find the popup element
            popup = page.locator('.esri-popup, .esri-feature-popup, [class*="popup"]').first
            
            # Extract all text content from popup
            popup_text = popup.inner_text()
            print(f"  Popup text: {popup_text[:200]}...")
            
            # Try to extract fields using regex patterns
            import re
            
            # State Parcel No
            state_parcel_match = re.search(r'State Parcel.*?:?\s*([\d\-\.]+)', popup_text, re.IGNORECASE)
            if state_parcel_match:
                data['state_parcel_no'] = state_parcel_match.group(1)
            
            # Owner Name
            owner_match = re.search(r'Owner.*?:?\s*([^\n]+)', popup_text, re.IGNORECASE)
            if owner_match:
                data['owner_name'] = owner_match.group(1).strip()
            
            # This is a fallback - popup may not have all the data we need
            # Return what we can extract
            return data
            
        except Exception as e:
            print(f"  Error extracting popup data: {e}")
            return None
    
    def _parse_address(self, address_text: str) -> Dict[str, str]:
        """
        Parse address text into components (street, city, state, zip)
        
        Handles formats like:
        - "123 Main St, City, ST 12345"
        - "123 Main St\nCity, ST 12345"
        - "123 Main St City ST 12345"
        - "26970 Pacific Terrace Dr, Mission Viejo, CA 92692"
        """
        import re
        
        result = {
            'street': '',
            'city': '',
            'state': '',
            'zip': ''
        }
        
        if not address_text:
            return result
        
        # Replace newlines with commas for easier parsing
        address_text = address_text.replace('\n', ', ').replace('\r', '')
        # Clean up multiple commas and spaces
        address_text = re.sub(r',\s*,', ',', address_text)
        address_text = re.sub(r'\s+', ' ', address_text).strip()
        
        # Try to extract ZIP code (5 digits or 5+4 format) - must be at the end or before end
        zip_match = re.search(r',?\s*(\d{5}(?:-\d{4})?)\s*$', address_text)
        if zip_match:
            result['zip'] = zip_match.group(1)
            # Remove ZIP from text (only from the end)
            address_text = address_text[:zip_match.start()].strip()
        
        # Try to extract state (2 letter code at the end)
        state_match = re.search(r',?\s*([A-Z]{2})\s*$', address_text)
        if state_match:
            result['state'] = state_match.group(1)
            # Remove state from text
            address_text = address_text[:state_match.start()].strip()
        
        # Split remaining text by comma
        parts = [p.strip() for p in address_text.split(',') if p.strip()]
        
        if len(parts) >= 2:
            # First part is street, last part is city
            result['street'] = parts[0]
            result['city'] = parts[-1]
        elif len(parts) == 1:
            # Only one part - assume it's the street
            result['street'] = parts[0]
        
        return result
    
    def _create_excel_template(self, county: str) -> openpyxl.Workbook:
        """Create Excel workbook with formatted headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{county} Parcels"
        
        # Header row
        headers = [
            'Parcel ID',
            'State Parcel No',
            'Owner Name',
            'Owner Address',
            'Owner City',
            'Owner State',
            'Owner Zip',
            'Parcel Address',
            'Parcel City',
            'Parcel State',
            'Parcel Zip',
            'Legal Description',
            'Latest Deed Date',
            'Document Number',
            'Deed Type',
            'Report Card Path',
            'Status'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_num, header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Parcel ID
        ws.column_dimensions['B'].width = 25  # State Parcel No
        ws.column_dimensions['C'].width = 30  # Owner Name
        ws.column_dimensions['D'].width = 30  # Owner Address
        ws.column_dimensions['E'].width = 20  # Owner City
        ws.column_dimensions['F'].width = 8   # Owner State
        ws.column_dimensions['G'].width = 12  # Owner Zip
        ws.column_dimensions['H'].width = 30  # Parcel Address
        ws.column_dimensions['I'].width = 20  # Parcel City
        ws.column_dimensions['J'].width = 8   # Parcel State
        ws.column_dimensions['K'].width = 12  # Parcel Zip
        ws.column_dimensions['L'].width = 50  # Legal Description
        ws.column_dimensions['M'].width = 15  # Deed Date
        ws.column_dimensions['N'].width = 20  # Document Number
        ws.column_dimensions['O'].width = 12  # Deed Type
        ws.column_dimensions['P'].width = 50  # Report Card Path
        ws.column_dimensions['Q'].width = 15  # Status
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        return wb
    
    def _safe_filename(self, filename: str) -> str:
        """Convert string to safe filename"""
        # Remove invalid characters
        filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
        # Limit length
        if len(filename) > 200:
            name, ext = os.path.splitext(filename)
            filename = name[:200-len(ext)] + ext
        return filename
    
    def _owner_filename_stub(self, owner_name: str) -> str:
        """
        Create a short filename stub from owner name
        Takes first 25 chars, removes special chars
        """
        if not owner_name or owner_name == 'Unknown':
            return 'Unknown'
        
        # Take first 25 characters
        stub = owner_name[:25]
        # Remove special characters, keep alphanumeric and spaces
        stub = re.sub(r'[^a-zA-Z0-9\s]', '', stub)
        # Replace spaces with underscores
        stub = stub.replace(' ', '_')
        # Remove multiple underscores
        stub = re.sub(r'_+', '_', stub)
        return stub.strip('_')
    
    def _download_prc(self, session, url: str, output_path: str):
        """Download PRC PDF with polite delay"""
        # Polite delay before download
        delay = random.uniform(*self.pdf_delay_range)
        time.sleep(delay)
        
        response = session.get(url, timeout=30)
        response.raise_for_status()
        
        with open(output_path, 'wb') as f:
            f.write(response.content)
