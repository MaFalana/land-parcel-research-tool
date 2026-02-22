"""
Beacon (Schneider) platform scraper

Handles scraping from beacon.schneidercorp.com portals across multiple counties.
Uses flexible selectors to adapt to county-specific variations in element IDs.
"""
from scrapers.base_scraper import BaseScraper
from typing import Dict, Callable, Optional
import os
import tempfile
import time
import random
import re
import requests
from urllib.parse import urlparse, parse_qs
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class BeaconScraper(BaseScraper):
    """Scraper for Beacon (Schneider) platform"""
    
    def __init__(self):
        super().__init__()
        self.platform_name = "Beacon"
        # Delay ranges (in seconds) - more conservative than ThinkGIS
        self.page_delay_range = (3, 7)  # 3-7 seconds between page requests
        self.pdf_delay_range = (2, 5)   # 2-5 seconds before PDF downloads
    
    def scrape_parcels(
        self,
        parcel_file_path: str,
        base_url: str,
        county: str,
        job_id: str,
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> Dict:
        """
        Scrape parcel data from Beacon portal
        
        Steps:
        1. Parse parcel IDs from input file
        2. Extract AppID and LayerID from base_url
        3. Open browser and navigate to Beacon
        4. For each parcel: search, extract data
        5. Save to Excel with progress tracking
        
        Args:
            parcel_file_path: Path to file with parcel IDs
            base_url: Beacon portal URL (contains AppID, LayerID)
            county: County name
            job_id: Job ID for output files
            progress_callback: Function to report progress
            
        Returns:
            Dict with excel_path and stats
        """
        # Parse parcel IDs
        parcel_ids = self.read_parcel_ids(parcel_file_path)
        total_parcels = len(parcel_ids)
        
        print(f"Beacon Scraper: Processing {total_parcels} parcels for {county} county")
        
        # Extract AppID and LayerID from URL
        url_params = self._extract_url_params(base_url)
        if not url_params['app_id'] or not url_params['layer_id']:
            raise ValueError(f"Could not extract AppID/LayerID from URL: {base_url}")
        
        print(f"Beacon Config: AppID={url_params['app_id']}, LayerID={url_params['layer_id']}")
        
        # Build search page URL (always use PageTypeID=2 for search)
        # Extract base domain from URL
        parsed = urlparse(base_url)
        search_url = f"{parsed.scheme}://{parsed.netloc}/Application.aspx?AppID={url_params['app_id']}&LayerID={url_params['layer_id']}&PageTypeID=2&PageID={url_params['page_id']}"
        print(f"Using search URL: {search_url}")
        
        # Create output directories
        output_dir = os.path.join(tempfile.gettempdir(), "parcel_jobs", job_id, "output")
        pdfs_dir = os.path.join(output_dir, "property_cards")
        os.makedirs(pdfs_dir, exist_ok=True)
        
        # Create Excel workbook
        excel_path = os.path.join(output_dir, f"{county}_beacon_data.xlsx")
        wb = self._create_excel_template(county)
        ws = wb.active
        
        # Track progress
        processed = 0
        failed = 0
        
        # Create requests session for PDF downloads
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/pdf,*/*",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        })
        
        # Launch browser (headless mode with args to avoid detection)
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox'
                ]
            )
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            try:
                # Navigate to Beacon portal (use search page URL)
                print(f"Navigating to: {search_url}")
                page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
                
                # Wait a bit for page to settle (don't wait for networkidle - Beacon has background activity)
                page.wait_for_timeout(3000)
                
                # Check if we got an error page
                if "Something went wrong" in page.content():
                    print("Error page detected, retrying...")
                    page.reload(wait_until="domcontentloaded")
                    page.wait_for_timeout(3000)
                
                # Click "Agree" or "Accept" button if present (terms and conditions)
                print("Checking for terms agreement...")
                agreement_clicked = False
                try:
                    # Try multiple variations of agree/accept buttons
                    agree_selectors = [
                        'text=Agree',
                        'text=Accept',
                        'button:has-text("Agree")',
                        'button:has-text("Accept")',
                        'input[value*="Agree"]',
                        'input[value*="Accept"]'
                    ]
                    
                    for selector in agree_selectors:
                        try:
                            button = page.locator(selector).first
                            if button.is_visible(timeout=3000):
                                print(f"Found agreement button: {selector}")
                                button.click()
                                agreement_clicked = True
                                # Wait for navigation/reload after clicking agree
                                page.wait_for_timeout(5000)
                                print("Clicked terms agreement button")
                                break
                        except:
                            continue
                except Exception as e:
                    print(f"No agreement button found or error: {e}")
                    pass  # No agree button, continue
                
                # If we clicked agreement, we might need to navigate back to search page
                if agreement_clicked:
                    print("Re-navigating to search page after agreement...")
                    page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
                    page.wait_for_timeout(3000)
                
                # Wait for search input to be available
                # Try multiple possible selectors (Beacon has different versions)
                print("Waiting for search input to be ready...")
                search_input_found = False
                search_selectors = [
                    'input#topSearchControl',  # New Beacon interface
                    'input[id*="txtParcelID"]',  # Old Beacon interface
                    'input[type="search"]'  # Generic search input
                ]
                
                for selector in search_selectors:
                    try:
                        page.wait_for_selector(selector, state="visible", timeout=5000)
                        print(f"Search input found: {selector}")
                        search_input_found = True
                        break
                    except:
                        continue
                
                if not search_input_found:
                    print(f"ERROR: Search input not found with any selector")
                    print(f"Current URL: {page.url}")
                    print(f"Page title: {page.title()}")
                    # Save screenshot for debugging
                    try:
                        screenshot_path = f"/tmp/beacon_debug_{job_id}.png"
                        page.screenshot(path=screenshot_path)
                        print(f"Screenshot saved to: {screenshot_path}")
                    except:
                        pass
                    raise Exception("Search input not found on page")
                
                # Process each parcel
                for idx, parcel_id in enumerate(parcel_ids, start=1):
                    row_num = idx + 2  # Data starts at row 3 (1-indexed)
                    
                    try:
                        print(f"Processing {idx}/{total_parcels}: {parcel_id}")
                        
                        # Search for parcel and extract data
                        parcel_data = self._search_parcel(page, parcel_id, search_url)
                        
                        if parcel_data:
                            # Write to Excel
                            ws.cell(row_num, 1, parcel_id)  # Column A: Parcel ID
                            ws.cell(row_num, 2, parcel_data.get('alternate_id', ''))  # Column B: Alternate ID
                            ws.cell(row_num, 3, parcel_data.get('owner_name', ''))  # Column C: Owner Name
                            ws.cell(row_num, 4, parcel_data.get('owner_address', ''))  # Column D: Owner Address
                            ws.cell(row_num, 5, parcel_data.get('owner_city', ''))  # Column E: Owner City
                            ws.cell(row_num, 6, parcel_data.get('owner_state', ''))  # Column F: Owner State
                            ws.cell(row_num, 7, parcel_data.get('owner_zip', ''))  # Column G: Owner Zip
                            ws.cell(row_num, 8, parcel_data.get('parcel_address', ''))  # Column H: Parcel Address
                            ws.cell(row_num, 9, parcel_data.get('parcel_city', ''))  # Column I: Parcel City
                            ws.cell(row_num, 10, parcel_data.get('parcel_state', ''))  # Column J: Parcel State
                            ws.cell(row_num, 11, parcel_data.get('parcel_zip', ''))  # Column K: Parcel Zip
                            ws.cell(row_num, 12, parcel_data.get('legal_description', ''))  # Column L: Legal Desc
                            ws.cell(row_num, 13, parcel_data.get('latest_deed_date', ''))  # Column M: Deed Date
                            ws.cell(row_num, 14, parcel_data.get('document_number', ''))  # Column N: Doc #
                            ws.cell(row_num, 15, parcel_data.get('deed_code', ''))  # Column O: Deed Type
                            
                            # Download PRC PDF if available
                            prc_path = ''
                            if parcel_data.get('prc_url'):
                                try:
                                    # Create filename: {owner_stub}_{parcel_id}.pdf
                                    owner_stub = self._owner_filename_stub(parcel_data.get('owner_name', 'Unknown'))
                                    pdf_filename = self._safe_filename(f"{owner_stub}_{parcel_id}.pdf")
                                    prc_full_path = os.path.join(pdfs_dir, pdf_filename)
                                    
                                    # Download PRC with polite delay
                                    self._download_prc(session, parcel_data['prc_url'], prc_full_path)
                                    prc_path = prc_full_path
                                    print(f"  ✓ Downloaded PRC: {pdf_filename}")
                                except Exception as e:
                                    print(f"  ✗ Failed to download PRC: {e}")
                                    import traceback
                                    traceback.print_exc()
                                    prc_path = f"ERROR: {str(e)[:50]}"
                            
                            ws.cell(row_num, 16, prc_path)  # Column P: Report Card Path
                            ws.cell(row_num, 17, 'SUCCESS')  # Column Q: Status
                            
                            processed += 1
                        else:
                            # Parcel not found
                            ws.cell(row_num, 1, parcel_id)
                            ws.cell(row_num, 17, 'NOT_FOUND')
                            failed += 1
                        
                        # Save progress every 10 parcels
                        if idx % 10 == 0:
                            wb.save(excel_path)
                            print(f"Progress saved: {processed} successful, {failed} failed")
                        
                        # Report progress
                        if progress_callback:
                            progress_callback(idx, total_parcels)
                        
                        # Polite delay between parcels (3-7 seconds)
                        delay = random.uniform(*self.page_delay_range)
                        time.sleep(delay)
                        
                        # Extra "thinking pause" every 15 parcels (10-15 seconds)
                        if idx % 15 == 0 and idx < total_parcels:
                            thinking_pause = random.uniform(10, 15)
                            print(f"  💭 Taking a thinking pause ({thinking_pause:.1f}s)...")
                            time.sleep(thinking_pause)
                        
                    except Exception as e:
                        print(f"Error processing {parcel_id}: {e}")
                        import traceback
                        traceback.print_exc()
                        ws.cell(row_num, 1, parcel_id)
                        ws.cell(row_num, 17, f'ERROR: {str(e)[:50]}')
                        failed += 1
                        continue
                
            finally:
                browser.close()
        
        # Final save
        wb.save(excel_path)
        
        print(f"\nBeacon Scraper: Complete!")
        print(f"  Processed: {processed}/{total_parcels}")
        print(f"  Failed: {failed}")
        print(f"  Excel: {excel_path}")
        print(f"  PDFs: {pdfs_dir}")
        
        return {
            "excel_path": excel_path,
            "pdfs_dir": pdfs_dir,
            "total": total_parcels,
            "processed": processed,
            "failed": failed
        }
    
    def _extract_url_params(self, url: str) -> Dict[str, str]:
        """Extract AppID, LayerID, PageTypeID, PageID from Beacon URL"""
        parsed = urlparse(url)
        params = parse_qs(parsed.query)
        
        return {
            'app_id': params.get('AppID', [''])[0],
            'layer_id': params.get('LayerID', [''])[0],
            'page_type_id': params.get('PageTypeID', ['1'])[0],
            'page_id': params.get('PageID', [''])[0]
        }
    
    def _parse_address(self, address_text: str) -> Dict[str, str]:
        """
        Parse address text into components (street, city, state, zip)
        
        Handles formats like:
        - "123 Main St, City, ST 12345"
        - "123 Main St\nCity, ST 12345"
        - "123 Main St City ST 12345"
        """
        import re
        
        result = {
            'street': '',
            'city': '',
            'state': '',
            'zip': ''
        }
        
        if not address_text:
            return result
        
        # Replace newlines with commas for easier parsing
        address_text = address_text.replace('\n', ', ').replace('\r', '')
        
        # Try to extract ZIP code (5 digits or 5+4 format)
        zip_match = re.search(r'\b(\d{5}(?:-\d{4})?)\b', address_text)
        if zip_match:
            result['zip'] = zip_match.group(1)
            # Remove ZIP from text
            address_text = address_text.replace(zip_match.group(0), '').strip()
        
        # Try to extract state (2 letter code before ZIP)
        state_match = re.search(r'\b([A-Z]{2})\s*,?\s*$', address_text)
        if state_match:
            result['state'] = state_match.group(1)
            # Remove state from text
            address_text = address_text[:state_match.start()].strip()
        
        # Split remaining text by comma
        parts = [p.strip() for p in address_text.split(',') if p.strip()]
        
        if len(parts) >= 2:
            # First part is street, last part is city
            result['street'] = parts[0]
            result['city'] = parts[-1]
        elif len(parts) == 1:
            # Only one part - assume it's the street
            result['street'] = parts[0]
        
        return result
    
    def _search_parcel(self, page, parcel_id: str, search_url: str) -> Optional[Dict]:
        """
        Search for a parcel and extract data using flexible selectors
        
        Args:
            page: Playwright page object
            parcel_id: Parcel ID to search for
            search_url: URL of the search page (PageTypeID=2)
        
        Returns dict with: owner_name, legal_description, latest_deed_date, 
                          document_number, deed_code, prc_url
        """
        try:
            # Find search input using flexible selectors (Beacon has different versions)
            search_selectors = [
                'input#topSearchControl',  # New Beacon interface
                'input[id*="txtParcelID"]',  # Old Beacon interface
                'input[type="search"]'  # Generic search input
            ]
            
            search_input = None
            for selector in search_selectors:
                try:
                    temp_input = page.locator(selector).first
                    if temp_input.is_visible(timeout=2000):
                        search_input = temp_input
                        print(f"  Using search input: {selector}")
                        break
                except:
                    continue
            
            if not search_input:
                print(f"ERROR: Cannot find search input for parcel {parcel_id}")
                print(f"Current URL: {page.url}")
                print(f"Expected search URL: {search_url}")
                # Check if we're on the right page
                if "PageTypeID=2" not in page.url:
                    print("WARNING: Not on search page! Navigating...")
                    page.goto(search_url, wait_until="domcontentloaded")
                    page.wait_for_timeout(3000)
                    # Try again
                    for selector in search_selectors:
                        try:
                            temp_input = page.locator(selector).first
                            if temp_input.is_visible(timeout=2000):
                                search_input = temp_input
                                break
                        except:
                            continue
                
                if not search_input:
                    raise Exception("Search input not found with any selector")
            
            # Clear and fill search box
            search_input.clear(timeout=5000)
            search_input.fill(parcel_id, timeout=5000)
            
            # Wait for autocomplete dropdown to appear (give it more time)
            page.wait_for_timeout(3000)
            
            # Look for autocomplete results (Twitter Typeahead)
            # The dropdown shows matching parcels - we need to click on one
            try:
                # Try to find the dropdown suggestion that matches our parcel
                # Typeahead creates suggestions with class 'tt-suggestion'
                # Try exact match first, then partial match
                suggestion = None
                
                # Try exact match
                try:
                    suggestion = page.locator(f'.tt-suggestion:has-text("{parcel_id}")').first
                    suggestion.wait_for(state="visible", timeout=3000)
                except:
                    # Try partial match - look for any suggestion containing part of the parcel ID
                    try:
                        # Get all suggestions
                        suggestions = page.locator('.tt-suggestion').all()
                        if suggestions:
                            # Click the first one (most relevant)
                            suggestion = suggestions[0]
                            print(f"  ⚠ Using first autocomplete suggestion (partial match)")
                    except:
                        pass
                
                if suggestion:
                    print(f"  ✓ Found autocomplete suggestion, clicking...")
                    suggestion.click()
                    
                    # Wait for navigation to property page (PageTypeID=4)
                    page.wait_for_timeout(5000)
                    
                    # Check if we navigated to property details page
                    if "PageTypeID=4" not in page.url:
                        print(f"  Warning: Expected PageTypeID=4, got: {page.url}")
                else:
                    raise Exception("No autocomplete suggestions found")
                    
            except Exception as e:
                # Autocomplete didn't work - parcel might not exist
                print(f"  ✗ Autocomplete failed: {str(e)[:100]}")
                page.goto(search_url)
                return None
            
            # Check if we got results or "no results" message
            # Beacon shows property details if found, or stays on search page if not
            try:
                # Try to find legal description (indicates we found the parcel)
                legal_desc_elem = page.locator('span[id*="lblLegalDescription"]').first
                legal_desc_elem.wait_for(timeout=3000)
            except:
                # Legal description not found = parcel not found
                print(f"Parcel {parcel_id} not found in Beacon")
                page.goto(search_url)  # Go back to search page
                return None
            
            # Extract data
            data = {}
            
            # Extract Parcel ID and Alternate ID (some counties use different formats)
            try:
                # Try to find parcel ID on the page
                parcel_id_elem = page.locator('span[id*="lblParcelID"], span[id*="lblParcel"]').first
                data['parcel_id_display'] = parcel_id_elem.inner_text(timeout=2000).strip()
            except:
                data['parcel_id_display'] = parcel_id
            
            try:
                # Try to find alternate ID
                alt_id_elem = page.locator('span[id*="lblAlternateID"], span[id*="lblAltID"], span[id*="AlternateID"]').first
                data['alternate_id'] = alt_id_elem.inner_text(timeout=2000).strip()
            except:
                data['alternate_id'] = ''
            
            # Owner name (from page title or owner section)
            try:
                # Try multiple selectors for owner name
                owner_selectors = [
                    'a[id*="lnkOwnerName"]',  # Link with owner name (most common in Beacon)
                    'span[id*="lblOwnerName"]',
                    'span[id*="lblOwner"]',
                    'span[id*="Owner1"]',
                    'td:has-text("Owner") + td span',
                    'th:has-text("Owner") + td span'
                ]
                
                owner_text = ''
                for selector in owner_selectors:
                    try:
                        elem = page.locator(selector).first
                        text = elem.inner_text(timeout=1000).strip()
                        if text and len(text) > 0:
                            # Make sure it's not an address (addresses usually have numbers at start)
                            # Owner names typically start with letters
                            if not text[0].isdigit():
                                owner_text = text
                                break
                    except:
                        continue
                
                data['owner_name'] = owner_text
            except Exception as e:
                data['owner_name'] = ''
            
            # Owner address (mailing address)
            try:
                # Try to find owner address - usually in a span with "Address" in the ID
                owner_addr_elem = page.locator('span[id*="lblOwnerAddress"], span[id*="OwnerAddress"]').first
                owner_addr_text = owner_addr_elem.inner_text(timeout=2000).strip()
                
                # Parse address into components
                # Format is usually: "123 Main St, City, ST 12345" or multiple lines
                addr_parts = self._parse_address(owner_addr_text)
                data['owner_address'] = addr_parts.get('street', '')
                data['owner_city'] = addr_parts.get('city', '')
                data['owner_state'] = addr_parts.get('state', '')
                data['owner_zip'] = addr_parts.get('zip', '')
            except:
                data['owner_address'] = ''
                data['owner_city'] = ''
                data['owner_state'] = ''
                data['owner_zip'] = ''
            
            # Parcel address (property location)
            try:
                # Try to find parcel/property address
                parcel_addr_elem = page.locator('span[id*="lblPropertyAddress"], span[id*="lblSitusAddress"], span[id*="lblLocation"]').first
                parcel_addr_text = parcel_addr_elem.inner_text(timeout=2000).strip()
                
                # Parse address into components
                addr_parts = self._parse_address(parcel_addr_text)
                data['parcel_address'] = addr_parts.get('street', '')
                data['parcel_city'] = addr_parts.get('city', '')
                data['parcel_state'] = addr_parts.get('state', '')
                data['parcel_zip'] = addr_parts.get('zip', '')
            except:
                data['parcel_address'] = ''
                data['parcel_city'] = ''
                data['parcel_state'] = ''
                data['parcel_zip'] = ''
            
            # Legal description
            try:
                data['legal_description'] = legal_desc_elem.inner_text().strip()
            except:
                data['legal_description'] = ''
            
            # Transfer history (latest deed)
            try:
                # Find transfer history table
                transfer_table = page.locator('table[id*="gvwTransferHistory"], table[id*="TransferHistory"]').first
                
                # Get first row (most recent transfer)
                first_row = transfer_table.locator('tbody tr').first
                
                # Date is usually in the first cell (th or td)
                date_cell = first_row.locator('th, td').first
                data['latest_deed_date'] = date_cell.inner_text(timeout=2000).strip()
                
                # Document number is usually in 3rd column
                doc_cell = first_row.locator('td').nth(2)
                data['document_number'] = doc_cell.inner_text(timeout=2000).strip()
                
                # Deed code might be in another column (WD, QC, etc.)
                # Try to find it
                try:
                    code_cell = first_row.locator('td').nth(1)
                    code_text = code_cell.inner_text(timeout=1000).strip()
                    # Check if it looks like a deed code (2-3 letters)
                    if len(code_text) <= 3 and code_text.isalpha():
                        data['deed_code'] = code_text
                    else:
                        data['deed_code'] = ''
                except:
                    data['deed_code'] = ''
                    
            except Exception as e:
                print(f"Could not extract transfer history: {e}")
                data['latest_deed_date'] = ''
                data['document_number'] = ''
                data['deed_code'] = ''
            
            # Extract PRC (Property Record Card) URL
            # Look for the most recent PRC link - there may be multiple years
            try:
                # Try to find PRC links - they usually have "Property Record Card" in the text
                prc_links = page.locator('a:has-text("Property Record Card")').all()
                
                if prc_links:
                    # If multiple PRCs, find the one with the most recent year
                    latest_prc = None
                    latest_year = 0
                    
                    for link in prc_links:
                        try:
                            link_text = link.inner_text()
                            # Extract year from text like "2024 Property Record Card (PDF)"
                            year_match = re.search(r'(\d{4})', link_text)
                            if year_match:
                                year = int(year_match.group(1))
                                if year > latest_year:
                                    latest_year = year
                                    latest_prc = link
                            elif not latest_prc:
                                # If no year found, use this as fallback
                                latest_prc = link
                        except:
                            continue
                    
                    if latest_prc:
                        prc_href = latest_prc.get_attribute('href')
                        
                        # Make absolute URL if relative
                        if prc_href.startswith('/'):
                            base_domain = f"{urlparse(search_url).scheme}://{urlparse(search_url).netloc}"
                            prc_href = base_domain + prc_href
                        elif not prc_href.startswith('http'):
                            base_domain = f"{urlparse(search_url).scheme}://{urlparse(search_url).netloc}"
                            prc_href = base_domain + '/' + prc_href
                        
                        data['prc_url'] = prc_href
                        if latest_year > 0:
                            print(f"  Found PRC URL: {latest_year} Property Record Card")
                        else:
                            print(f"  Found PRC URL: {prc_href[:80]}...")
                    else:
                        data['prc_url'] = None
                else:
                    # Fallback: try generic PDF links
                    pdf_links = page.locator('a[href*=".pdf"]').all()
                    if pdf_links:
                        prc_href = pdf_links[0].get_attribute('href')
                        if prc_href.startswith('/'):
                            base_domain = f"{urlparse(search_url).scheme}://{urlparse(search_url).netloc}"
                            prc_href = base_domain + prc_href
                        elif not prc_href.startswith('http'):
                            base_domain = f"{urlparse(search_url).scheme}://{urlparse(search_url).netloc}"
                            prc_href = base_domain + '/' + prc_href
                        data['prc_url'] = prc_href
                        print(f"  Found PDF URL: {prc_href[:80]}...")
                    else:
                        data['prc_url'] = None
            except Exception as e:
                print(f"Could not find PRC URL: {e}")
                data['prc_url'] = None
            
            # Go back to search page for next parcel
            page.goto(search_url, wait_until="domcontentloaded")
            page.wait_for_timeout(2000)
            
            return data
            
        except Exception as e:
            print(f"Error searching for parcel {parcel_id}: {e}")
            import traceback
            traceback.print_exc()
            # Try to recover by going back to search page
            try:
                page.goto(search_url, wait_until="domcontentloaded")
                page.wait_for_timeout(2000)
            except:
                pass
            return None
    
    def _create_excel_template(self, county: str) -> openpyxl.Workbook:
        """Create Excel workbook with formatted headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{county} Parcels"
        
        # Header row (row 1)
        headers = [
            'Parcel ID',
            'Alternate ID',
            'Owner Name',
            'Owner Address',
            'Owner City',
            'Owner State',
            'Owner Zip',
            'Parcel Address',
            'Parcel City',
            'Parcel State',
            'Parcel Zip',
            'Legal Description',
            'Latest Deed Date',
            'Document Number',
            'Deed Type',
            'Report Card Path',
            'Status'
        ]
        
        # Style for headers
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_num, header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Parcel ID
        ws.column_dimensions['B'].width = 25  # Alternate ID
        ws.column_dimensions['C'].width = 30  # Owner Name
        ws.column_dimensions['D'].width = 30  # Owner Address
        ws.column_dimensions['E'].width = 20  # Owner City
        ws.column_dimensions['F'].width = 8   # Owner State
        ws.column_dimensions['G'].width = 12  # Owner Zip
        ws.column_dimensions['H'].width = 30  # Parcel Address
        ws.column_dimensions['I'].width = 20  # Parcel City
        ws.column_dimensions['J'].width = 8   # Parcel State
        ws.column_dimensions['K'].width = 12  # Parcel Zip
        ws.column_dimensions['L'].width = 50  # Legal Description
        ws.column_dimensions['M'].width = 15  # Deed Date
        ws.column_dimensions['N'].width = 20  # Document Number
        ws.column_dimensions['O'].width = 12  # Deed Type
        ws.column_dimensions['P'].width = 50  # Report Card Path
        ws.column_dimensions['Q'].width = 15  # Status
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        return wb

    def _safe_filename(self, filename: str) -> str:
        """Make a filename safe for filesystem"""
        # Remove or replace unsafe characters
        safe = re.sub(r'[<>:"/\\|?*]', '_', filename)
        # Remove leading/trailing spaces and dots
        safe = safe.strip('. ')
        # Limit length
        if len(safe) > 200:
            name, ext = os.path.splitext(safe)
            safe = name[:196] + ext
        return safe
    
    def _owner_filename_stub(self, owner_name: str) -> str:
        """Create a short filename stub from owner name"""
        if not owner_name:
            return "Unknown"
        
        # Take first 30 chars, remove special chars
        stub = owner_name[:30]
        stub = re.sub(r'[^a-zA-Z0-9\s]', '', stub)
        stub = stub.strip().replace(' ', '_')
        
        return stub if stub else "Unknown"
    
    def _download_prc(self, session, url: str, output_path: str):
        """Download Property Record Card PDF with polite delay"""
        # Check if already downloaded
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            return output_path
        
        # Polite delay before PDF download
        delay = random.uniform(*self.pdf_delay_range)
        time.sleep(delay)
        
        # Download PDF
        response = session.get(url, timeout=45)
        response.raise_for_status()
        
        # Save to file
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'wb') as f:
            f.write(response.content)
        
        return output_path
